import os
import re
import json
import secrets
import unicodedata
import uuid
import logging
from threading import Lock
from io import BytesIO
import csv
import tempfile
from datetime import datetime, timedelta
from time import perf_counter
from .crud import now_sv

from dotenv import load_dotenv
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile, status
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import extract, func, or_
from sqlalchemy.orm import Session, joinedload
from starlette.middleware.sessions import SessionMiddleware
from urllib.parse import quote_plus

from .auth import (
    authenticate_user,
    get_current_user,
    hash_password,
    login_user,
    logout_user,
    verify_password,
)
from .catalog_types import catalog_item_allows_decimal, classify_catalog_item_type
from .crud import (
    agregar_item_db,
    calcular_ingreso_pk_bodega,
    calcular_retorno_esperado,
    ejecutar_liquidacion,
    marcar_liquidada_en_prokey,
    puede_liquidar,
    puede_preparar,
    crear_requisicion_db,
    parse_items_from_form,
    puede_aprobar,
    puede_entregar,
    transicionar_requisicion,
    validar_liquidacion_item,
)
from .database import get_db, run_migrations
from .database import (
    create_backup_archive,
    get_backup_directory,
    is_sqlite_database,
    list_backup_archives,
    resolve_backup_archive,
    restore_backup_archive,
)
from .logging_utils import clear_request_id, set_request_id, setup_logging
from .models import CatalogoItem, Item, Requisicion, Usuario

load_dotenv()
setup_logging()
logger = logging.getLogger(__name__)

SECRET_KEY = os.getenv("SECRET_KEY", "change-me")

app = FastAPI(title=os.getenv("APP_NAME", "Sistema de Requisiciones MVP"))
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

DEPARTAMENTOS_VALIDOS = ["Cuentas", "Ventas", "Bodega", "Admon", "Logistica"]
CATALOGO_HEADERS = {"nombre", "item", "producto", "descripcion"}
ROLES_VALIDOS = ["user", "logistica", "aprobador", "bodega", "jefe_bodega", "admin", "tecnico"]
PASSWORD_MIN_LENGTH = 8
USUARIOS_IMPORT_HEADERS = {"nombre", "puesto"}
TEMP_IMPORT_PASSWORD = "Temp@2026"
TEMP_IMPORT_PIN = "1234"
MOTIVOS_REQUISICION = [
    "Queja Fragancia",
    "Otros",
    "Demostración",
    "Queja Eq. Dañado",
    "Servicio No Programado",
    "R1E",
    "Queja Mal Estado",
    "Restauracion Eq. Calidad",
    "Reponer KIT a Tecnico",
    "Cambio de Fragancia",
    "Servicio pendiente",
]
MAINTENANCE_STATE = {"active": False, "reason": None}
BACKUP_OPERATION_LOCK = Lock()


def build_catalog_payload(items: list[CatalogoItem]) -> list[dict[str, object]]:
    return [
        {
            "nombre": item.nombre,
            "tipo_item": item.tipo_item,
            "permite_decimal": bool(item.permite_decimal),
        }
        for item in items
    ]
PUESTO_MAP = {
    "AUXILIAR DE BODEGA": ("bodega", "Bodega"),
    "TECNICO DE SERVICIO": ("tecnico", "Logistica"),
    "EJECUTIVO DE CUENTAS": ("user", "Cuentas"),
    "EJECUTIVA DE CUENTAS": ("user", "Cuentas"),
    "EJECUTIVO DE VENTAS": ("user", "Ventas"),
    "ASISTENTE ADMINISTRATIVO": ("user", "Admon"),
    "GERENTE GENERAL": ("aprobador", "Admon"),
    "JEFE ADMINISTRATIVO": ("aprobador", "Admon"),
    "JEFE DE BODEGA": ("aprobador", "Admon"),
    "JEFE DE OPERACIONES": ("aprobador", "Admon"),
    "JEFE DE LOGISTICA": ("aprobador", "Admon"),
}


def get_client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    client = request.client
    return client.host if client else "-"


def format_file_size(size_bytes: int) -> str:
    size = float(size_bytes or 0)
    for unit in ("B", "KB", "MB", "GB"):
        if size < 1024 or unit == "GB":
            if unit == "B":
                return f"{int(size)} {unit}"
            return f"{size:.1f} {unit}"
        size /= 1024
    return f"{size_bytes} B"


def maintenance_response(request: Request) -> Response:
    message = MAINTENANCE_STATE.get("reason") or "La aplicacion esta restaurando un respaldo. Intenta nuevamente en unos segundos."
    if request.url.path.startswith("/api/"):
        return JSONResponse(status_code=503, content={"detail": message})
    html = (
        "<!DOCTYPE html><html lang='es'><head><meta charset='UTF-8'>"
        "<meta name='viewport' content='width=device-width, initial-scale=1.0'>"
        "<title>Mantenimiento</title></head>"
        "<body style=\"font-family: Montserrat, sans-serif; background:#071a2e; color:#e7f0ff; display:flex; "
        "align-items:center; justify-content:center; min-height:100vh; margin:0;\">"
        "<div style=\"max-width:560px; padding:1.5rem; border:1px solid rgba(255,255,255,.12); "
        "border-radius:16px; background:rgba(11,31,52,.96); box-shadow:0 12px 30px rgba(0,0,0,.28);\">"
        "<h1 style=\"margin:0 0 .75rem; font-size:1.4rem; color:#98d3ff;\">Restaurando respaldo</h1>"
        f"<p style=\"margin:0; line-height:1.55;\">{message}</p>"
        "</div></body></html>"
    )
    return HTMLResponse(status_code=503, content=html)


def build_backup_rows() -> list[dict[str, object]]:
    rows = list_backup_archives()
    for row in rows:
        row["size_human"] = format_file_size(int(row.get("size_bytes") or 0))
    return rows


def get_session_user_id(request: Request) -> int | None:
    try:
        return request.session.get("user_id")
    except Exception:
        return None


@app.middleware("http")
async def request_logging_middleware(request: Request, call_next):
    request_id = request.headers.get("X-Request-ID") or uuid.uuid4().hex[:12]
    set_request_id(request_id)
    start = perf_counter()
    client_ip = get_client_ip(request)
    user_id = get_session_user_id(request)
    try:
        if MAINTENANCE_STATE["active"] and not request.url.path.startswith("/static"):
            response = maintenance_response(request)
            response.headers["X-Request-ID"] = request_id
            logger.warning(
                "maintenance_blocked_request",
                extra={
                    "event": "maintenance_blocked_request",
                    "method": request.method,
                    "path": request.url.path,
                    "status_code": response.status_code,
                    "client_ip": client_ip,
                    "user_id": user_id,
                },
            )
            return response
        response = await call_next(request)
        duration_ms = round((perf_counter() - start) * 1000, 2)
        response.headers["X-Request-ID"] = request_id
        if not request.url.path.startswith("/static"):
            logger.info(
                "request_completed",
                extra={
                    "event": "request_completed",
                    "method": request.method,
                    "path": request.url.path,
                    "status_code": response.status_code,
                    "duration_ms": duration_ms,
                    "client_ip": client_ip,
                    "user_id": user_id,
                },
            )
        return response
    except Exception:
        duration_ms = round((perf_counter() - start) * 1000, 2)
        logger.exception(
            "request_failed",
            extra={
                "event": "request_failed",
                "method": request.method,
                "path": request.url.path,
                "duration_ms": duration_ms,
                "client_ip": client_ip,
                "user_id": user_id,
            },
        )
        raise
    finally:
        clear_request_id()


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    log_level = logging.ERROR if exc.status_code >= 500 else logging.WARNING
    logger.log(
        log_level,
        "http_exception",
        extra={
            "event": "http_exception",
            "method": request.method,
            "path": request.url.path,
            "status_code": exc.status_code,
            "client_ip": get_client_ip(request),
            "reason": str(exc.detail),
            "user_id": get_session_user_id(request),
        },
    )
    if exc.status_code == status.HTTP_401_UNAUTHORIZED:
        if request.url.path.startswith("/api/"):
            return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})
        return RedirectResponse(url="/login", status_code=303)
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})


@app.on_event("startup")
def startup_migrations() -> None:
    run_migrations()


def normalize_catalog_name(value: str) -> str:
    return " ".join(value.split()).strip().casefold()


def normalize_text(value: str) -> str:
    base = unicodedata.normalize("NFD", value or "")
    without_marks = "".join(ch for ch in base if unicodedata.category(ch) != "Mn")
    return " ".join(without_marks.split()).strip()


def normalize_person_name(value: str) -> str:
    return normalize_text(value).upper()


def normalize_puesto(value: str) -> str:
    return normalize_text(value).upper()


def build_username_base(nombre: str) -> str:
    parts = [re.sub(r"[^a-z0-9]", "", p) for p in normalize_text(nombre).lower().split()]
    parts = [p for p in parts if p]
    if not parts:
        return "usuario"
    if len(parts) == 1:
        return parts[0]
    primer_apellido = parts[-2] if len(parts) >= 3 else parts[1]
    return f"{parts[0][0]}{primer_apellido}"


def pick_unique_username(base: str, taken: set[str]) -> str:
    candidate = base
    suffix = 1
    while candidate in taken:
        suffix += 1
        candidate = f"{base}{suffix}"
    taken.add(candidate)
    return candidate


def _extract_names_from_rows(rows: list[list[str]]) -> list[str]:
    if not rows:
        return []
    header_idx = -1
    first = [c.strip().casefold() for c in rows[0]]
    for idx, value in enumerate(first):
        if value in CATALOGO_HEADERS:
            header_idx = idx
            break

    start = 1 if header_idx >= 0 else 0
    name_idx = header_idx if header_idx >= 0 else 0

    names: list[str] = []
    seen: set[str] = set()
    for row in rows[start:]:
        if name_idx >= len(row):
            continue
        name = " ".join(str(row[name_idx]).split()).strip()
        if len(name) < 2:
            continue
        key = normalize_catalog_name(name)
        if key in seen:
            continue
        seen.add(key)
        names.append(name)
    return names


def _parse_catalog_csv(raw: bytes) -> list[str]:
    text = ""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if not text:
        raise ValueError("No se pudo leer el CSV")
    reader = csv.reader(text.splitlines())
    rows = [[str(cell).strip() for cell in row] for row in reader if any(str(cell).strip() for cell in row)]
    return _extract_names_from_rows(rows)


def _parse_catalog_xlsx(raw: bytes) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ValueError("Falta dependencia openpyxl para importar XLSX") from exc

    wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            values = ["" if value is None else str(value).strip() for value in row]
            if any(values):
                rows.append(values)
        return _extract_names_from_rows(rows)
    finally:
        wb.close()


def _parse_users_rows(raw: bytes, filename: str) -> list[dict[str, str]]:
    ext = os.path.splitext(filename.lower())[1]
    rows: list[list[str]] = []
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise ValueError("Falta dependencia openpyxl para importar XLSX") from exc
        wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                values = ["" if value is None else str(value).strip() for value in row]
                if any(values):
                    rows.append(values)
        finally:
            wb.close()
    elif ext == ".csv":
        text = ""
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if not text:
            raise ValueError("No se pudo leer el CSV")
        reader = csv.reader(text.splitlines())
        rows = [[str(cell).strip() for cell in row] for row in reader if any(str(cell).strip() for cell in row)]
    else:
        raise ValueError("Formato no soportado. Usa XLSX o CSV")

    if not rows:
        raise ValueError("El archivo no contiene filas")

    headers = [normalize_text(col).lower() for col in rows[0]]
    try:
        idx_nombre = headers.index("nombre")
        idx_puesto = headers.index("puesto")
    except ValueError as exc:
        raise ValueError("El archivo debe incluir columnas NOMBRE y PUESTO") from exc

    parsed: list[dict[str, str]] = []
    for line_num, row in enumerate(rows[1:], start=2):
        nombre = row[idx_nombre].strip() if idx_nombre < len(row) else ""
        puesto = row[idx_puesto].strip() if idx_puesto < len(row) else ""
        if not nombre and not puesto:
            continue
        parsed.append({"linea": str(line_num), "nombre": nombre, "puesto": puesto})
    return parsed


def template_context(request: Request, current_user: Usuario | None = None, **kwargs: object) -> dict[str, object]:
    return {"request": request, "current_user": current_user, **kwargs}


def format_datetime(value: datetime | str | None) -> str:
    if not value:
        return "-"
    if isinstance(value, str):
        cleaned = value.replace("T", " ").split(".")[0].strip()
        return cleaned if cleaned else "-"
    return value.strftime("%Y-%m-%d %H:%M:%S")


templates.env.filters["fmt_dt"] = format_datetime


def attach_catalog_item_defaults(items: list, db: Session) -> None:
    catalog_map = {
        normalize_catalog_name(catalog_item.nombre): {
            "default_mode": (catalog_item.tipo_item or classify_catalog_item_type(catalog_item.nombre)),
            "permite_decimal": bool(catalog_item.permite_decimal),
        }
        for catalog_item in db.query(CatalogoItem).all()
    }
    for item in items:
        catalog_defaults = catalog_map.get(normalize_catalog_name(item.descripcion), {})
        item.default_mode = catalog_defaults.get("default_mode")
        item.permite_decimal = bool(catalog_defaults.get("permite_decimal", False))


def normalize_optional_text(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = value.strip()
    return cleaned or None


def can_view_requisicion(req: Requisicion, current_user: Usuario) -> bool:
    return (
        current_user.rol == "admin"
        or current_user.id == req.solicitante_id
        or current_user.rol in ["aprobador", "jefe_bodega", "logistica"]
        or (
            current_user.rol == "bodega"
            and req.estado in ["aprobada", "preparado", "entregada", "liquidada", "liquidada_en_prokey"]
        )
    )


def normalize_contexto_operacion(value: str | None) -> str | None:
    cleaned = normalize_optional_text(value)
    if cleaned is None:
        return None
    lowered = cleaned.lower()
    if lowered not in ("reposicion", "instalacion_inicial"):
        return None
    return lowered


def redirect_with_message(url: str, message: str, level: str = "success") -> RedirectResponse:
    safe_msg = quote_plus(message)
    safe_level = quote_plus(level)
    sep = "&" if "?" in url else "?"
    return RedirectResponse(url=f"{url}{sep}msg={safe_msg}&type={safe_level}", status_code=303)


def puede_editar_prokey_ref(req: Requisicion, current_user: Usuario) -> bool:
    return req.estado == "liquidada" and (
        current_user.rol in ("admin", "logistica") or req.solicitante_id == current_user.id
    )


def es_bodega_plano(current_user: Usuario) -> bool:
    return current_user.rol == "bodega"


def puede_ver_todas_las_requisiciones(current_user: Usuario) -> bool:
    return current_user.rol == "logistica"


def ensure_all_requests_access(current_user: Usuario) -> None:
    if current_user.rol not in ["admin", "aprobador", "jefe_bodega", "logistica"]:
        raise HTTPException(status_code=403, detail="No autorizado")


def redirect_if_bodega_plain_accesses_own_requests(current_user: Usuario) -> RedirectResponse | None:
    if es_bodega_plano(current_user):
        return redirect_with_message(
            "/bodega",
            "El rol bodega solo puede operar desde la vista de bodega",
            "warning",
        )
    return None


def ensure_admin(current_user: Usuario) -> None:
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")


def build_home_cards(current_user: Usuario, db: Session) -> list[dict[str, object]]:
    ahora = now_sv()
    inicio_mes = datetime(ahora.year, ahora.month, 1)
    mis_query = db.query(Requisicion).filter(Requisicion.solicitante_id == current_user.id)

    mis_requisiciones = mis_query.count()
    mis_abiertas = mis_query.filter(
        Requisicion.estado.in_(["pendiente", "aprobada", "preparado", "entregada", "liquidada"]),
        or_(Requisicion.delivery_result.is_(None), Requisicion.delivery_result != "no_entregada"),
    ).count()
    mis_cerradas = mis_query.filter(
        or_(
            Requisicion.estado == "liquidada_en_prokey",
            Requisicion.estado == "no_entregada",
            Requisicion.delivery_result == "no_entregada",
        )
    ).count()
    mis_creadas_mes = mis_query.filter(Requisicion.created_at >= inicio_mes).count()

    pendientes_aprobar = db.query(Requisicion).filter(Requisicion.estado == "pendiente").count()
    pendientes_entregar = db.query(Requisicion).filter(Requisicion.estado.in_(["aprobada", "preparado"])).count()
    preparadas = db.query(Requisicion).filter(Requisicion.estado == "preparado").count()
    pendientes_liquidar = db.query(Requisicion).filter(Requisicion.estado == "entregada").count()
    liquidadas = db.query(Requisicion).filter(Requisicion.estado == "liquidada").count()
    liquidadas_en_prokey = db.query(Requisicion).filter(Requisicion.estado == "liquidada_en_prokey").count()
    rechazadas = db.query(Requisicion).filter(Requisicion.estado == "rechazada").count()
    no_entregadas = db.query(Requisicion).filter(
        or_(Requisicion.estado == "no_entregada", Requisicion.delivery_result == "no_entregada")
    ).count()
    pendientes_ref_prokey = db.query(Requisicion).filter(
        Requisicion.estado == "liquidada",
        or_(Requisicion.prokey_ref.is_(None), Requisicion.prokey_ref == ""),
    ).count()
    todas_requisiciones = db.query(Requisicion).count()

    cards_by_role = {
        "user": [
            {"label": "Todas Mis Requisiciones", "value": mis_requisiciones, "href": "/mis-requisiciones", "icon": "list"},
            {
                "label": "Requisiciones Pendientes",
                "value": mis_abiertas,
                "href": "/mis-requisiciones?estado=abiertas",
                "icon": "pending",
            },
            {"label": "Requisiciones Finalizadas", "value": mis_cerradas, "href": "/mis-requisiciones?estado=cerradas", "icon": "closed"},
            {"label": "Creadas Este Mes", "value": mis_creadas_mes, "href": "/mis-requisiciones", "icon": "month"},
        ],
        "logistica": [
            {"label": "Todas Mis Requisiciones", "value": mis_requisiciones, "href": "/mis-requisiciones", "icon": "list"},
            {
                "label": "Requisiciones Pendientes",
                "value": mis_abiertas,
                "href": "/mis-requisiciones?estado=abiertas",
                "icon": "pending",
            },
            {"label": "Requisiciones Finalizadas", "value": mis_cerradas, "href": "/mis-requisiciones?estado=cerradas", "icon": "closed"},
            {"label": "Todas las Requisiciones", "value": todas_requisiciones, "href": "/todas-requisiciones", "icon": "all"},
            {
                "label": "Pendientes de Referencia Prokey",
                "value": pendientes_ref_prokey,
                "href": "/todas-requisiciones?estado=liquidada",
                "icon": "prokey_pending",
            },
            {
                "label": "Liquidadas en Prokey",
                "value": liquidadas_en_prokey,
                "href": "/todas-requisiciones?estado=liquidada_en_prokey",
                "icon": "prokey_done",
            },
        ],
        "aprobador": [
            {"label": "Pendientes por Aprobar", "value": pendientes_aprobar, "href": "/aprobar", "icon": "approve"},
            {
                "label": "Pendientes de Entregar",
                "value": pendientes_entregar,
                "href": "/todas-requisiciones?estado=pendiente_entregar",
                "icon": "deliver",
            },
            {
                "label": "Pendientes de Liquidar",
                "value": pendientes_liquidar,
                "href": "/todas-requisiciones?estado=entregada",
                "icon": "liquidate",
            },
            {
                "label": "Requisiciones Rechazadas",
                "value": rechazadas,
                "href": "/todas-requisiciones?estado=rechazada",
                "icon": "rejected",
            },
        ],
        "bodega": [
            {"label": "Pendientes de Procesar", "value": pendientes_entregar, "href": "/bodega", "icon": "warehouse"},
            {
                "label": "Pendientes de Liquidar",
                "value": pendientes_liquidar,
                "href": "/bodega",
                "icon": "liquidate",
            },
            {"label": "Liquidadas", "value": liquidadas, "href": "/bodega?vista=historial", "icon": "closed"},
            {
                "label": "Liquidadas en Prokey",
                "value": liquidadas_en_prokey,
                "href": "/bodega?vista=historial",
                "icon": "prokey_done",
            },
        ],
        "jefe_bodega": [
            {"label": "Pendientes por Aprobar", "value": pendientes_aprobar, "href": "/aprobar", "icon": "approve"},
            {"label": "Pendientes de Procesar", "value": pendientes_entregar, "href": "/bodega", "icon": "warehouse"},
            {
                "label": "Pendientes de Liquidar",
                "value": pendientes_liquidar,
                "href": "/bodega",
                "icon": "liquidate",
            },
            {
                "label": "Liquidadas en Prokey",
                "value": liquidadas_en_prokey,
                "href": "/bodega?vista=historial",
                "icon": "prokey_done",
            },
        ],
        "admin": [
            {"label": "Todas Mis Requisiciones", "value": mis_requisiciones, "href": "/mis-requisiciones", "icon": "list"},
            {"label": "Pendientes por Aprobar", "value": pendientes_aprobar, "href": "/aprobar", "icon": "approve"},
            {
                "label": "Pendientes de Entregar",
                "value": pendientes_entregar,
                "href": "/todas-requisiciones?estado=pendiente_entregar",
                "icon": "deliver",
            },
            {
                "label": "Pendientes de Liquidar",
                "value": pendientes_liquidar,
                "href": "/todas-requisiciones?estado=entregada",
                "icon": "liquidate",
            },
            {"label": "Liquidadas", "value": liquidadas, "href": "/todas-requisiciones?estado=liquidada", "icon": "closed"},
            {
                "label": "Liquidadas en Prokey",
                "value": liquidadas_en_prokey,
                "href": "/todas-requisiciones?estado=liquidada_en_prokey",
                "icon": "prokey_done",
            },
        ],
    }
    return cards_by_role.get(current_user.rol, cards_by_role["user"])


def build_home_actions(current_user: Usuario) -> list[dict[str, str]]:
    actions: list[dict[str, str]] = []
    if current_user.rol != "bodega":
        actions.append({"label": "Nueva Requisición", "href": "/crear", "icon": "new"})

    if current_user.rol in ["admin", "aprobador", "jefe_bodega", "logistica"]:
        actions.append({"label": "Todas las Requisiciones", "href": "/todas-requisiciones", "icon": "search"})
    elif current_user.rol == "bodega":
        actions.append({"label": "Bodega", "href": "/bodega", "icon": "warehouse"})
    else:
        actions.append({"label": "Mis Requisiciones", "href": "/mis-requisiciones", "icon": "search"})

    if current_user.rol in ["admin", "aprobador", "jefe_bodega"]:
        actions.append({"label": "Aprobar", "href": "/aprobar", "icon": "approve"})
        actions.append({"label": "Monitor de Actividad", "href": "/monitor", "icon": "monitor"})

    if current_user.rol in ["admin", "bodega", "jefe_bodega"]:
        actions.append({"label": "Bodega", "href": "/bodega", "icon": "warehouse"})

    return actions


def build_home_user_status_chart(current_user: Usuario, db: Session) -> dict[str, object] | None:
    if current_user.rol != "user":
        return None

    mis_query = db.query(Requisicion).filter(Requisicion.solicitante_id == current_user.id)
    total = mis_query.count()
    segmentos_raw = [
        {
            "label": "Pendiente de aprobación",
            "value": mis_query.filter(Requisicion.estado == "pendiente").count(),
            "tone": "pending",
        },
        {
            "label": "En proceso",
            "value": mis_query.filter(Requisicion.estado.in_(["aprobada", "preparado"])).count(),
            "tone": "process",
        },
        {
            "label": "Pendiente de cierre",
            "value": mis_query.filter(
                Requisicion.estado.in_(["entregada", "liquidada"]),
                or_(Requisicion.delivery_result.is_(None), Requisicion.delivery_result != "no_entregada"),
            ).count(),
            "tone": "closure",
        },
        {
            "label": "Rechazada",
            "value": mis_query.filter(Requisicion.estado == "rechazada").count(),
            "tone": "rejected",
        },
        {
            "label": "Finalizada",
            "value": mis_query.filter(
                or_(
                    Requisicion.estado == "liquidada_en_prokey",
                    Requisicion.estado == "no_entregada",
                    Requisicion.delivery_result == "no_entregada",
                )
            ).count(),
            "tone": "finalized",
        },
    ]
    segmentos = []
    for segmento in segmentos_raw:
        porcentaje = round((segmento["value"] * 100 / total), 1) if total else 0
        segmentos.append(
            {
                **segmento,
                "percentage": porcentaje,
                "width_pct": max(porcentaje, 3) if segmento["value"] > 0 and total else 0,
            }
        )

    return {
        "total": total,
        "segments": segmentos,
        "has_data": total > 0,
    }


def build_home_user_monthly_chart(current_user: Usuario, db: Session) -> dict[str, object] | None:
    if current_user.rol != "user":
        return None

    ahora = now_sv()
    months = []
    year = ahora.year
    month = ahora.month
    for _ in range(6):
        months.append((year, month))
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    months.reverse()

    labels = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    bars = []
    max_value = 1
    for year_value, month_value in months:
        start = datetime(year_value, month_value, 1)
        if month_value == 12:
            end = datetime(year_value + 1, 1, 1)
        else:
            end = datetime(year_value, month_value + 1, 1)
        count = (
            db.query(Requisicion)
            .filter(
                Requisicion.solicitante_id == current_user.id,
                Requisicion.created_at >= start,
                Requisicion.created_at < end,
            )
            .count()
        )
        max_value = max(max_value, count)
        bars.append(
            {
                "label": f"{labels[month_value - 1]} {str(year_value)[-2:]}",
                "value": count,
            }
        )

    for bar in bars:
        bar["height_pct"] = max(round((bar["value"] * 100) / max_value, 1), 8) if bar["value"] > 0 else 0

    return {
        "bars": bars,
        "has_data": any(bar["value"] > 0 for bar in bars),
        "max_value": max_value,
    }


def build_home_user_closure_chart(current_user: Usuario, db: Session) -> dict[str, object] | None:
    if current_user.rol != "user":
        return None

    closed_requests = (
        db.query(Requisicion)
        .filter(
            Requisicion.solicitante_id == current_user.id,
            or_(
                Requisicion.estado == "liquidada_en_prokey",
                Requisicion.estado == "no_entregada",
                Requisicion.delivery_result == "no_entregada",
            ),
        )
        .all()
    )

    buckets = [
        {"label": "0-1 días", "value": 0, "tone": "fast"},
        {"label": "2-3 días", "value": 0, "tone": "medium"},
        {"label": "4-7 días", "value": 0, "tone": "slow"},
        {"label": "8+ días", "value": 0, "tone": "very-slow"},
    ]

    for req in closed_requests:
        end_at = req.prokey_liquidada_at or req.liquidated_at or req.delivered_at or req.updated_at or req.created_at
        start_at = req.created_at
        if not start_at or not end_at:
            continue
        delta_days = max((end_at - start_at).total_seconds() / 86400, 0)
        if delta_days <= 1:
            buckets[0]["value"] += 1
        elif delta_days <= 3:
            buckets[1]["value"] += 1
        elif delta_days <= 7:
            buckets[2]["value"] += 1
        else:
            buckets[3]["value"] += 1

    total = sum(bucket["value"] for bucket in buckets)
    max_value = max(1, *(bucket["value"] for bucket in buckets))
    for bucket in buckets:
        bucket["height_pct"] = max(round((bucket["value"] * 100) / max_value, 1), 10) if bucket["value"] > 0 else 0
        bucket["percentage"] = round((bucket["value"] * 100) / total, 1) if total else 0

    return {
        "bars": buckets,
        "has_data": total > 0,
        "total": total,
    }


def build_home_bodega_status_chart(current_user: Usuario, db: Session) -> dict[str, object] | None:
    if current_user.rol != "bodega":
        return None

    total = db.query(Requisicion).count()
    segmentos_raw = [
        {
            "label": "Pendientes de Procesar",
            "value": db.query(Requisicion).filter(Requisicion.estado.in_(["aprobada", "preparado"])).count(),
            "tone": "process",
        },
        {
            "label": "Pendientes de Liquidar",
            "value": db.query(Requisicion).filter(Requisicion.estado == "entregada").count(),
            "tone": "closure",
        },
        {
            "label": "Liquidadas",
            "value": db.query(Requisicion).filter(Requisicion.estado == "liquidada").count(),
            "tone": "pending",
        },
        {
            "label": "Liquidadas en Prokey",
            "value": db.query(Requisicion).filter(Requisicion.estado == "liquidada_en_prokey").count(),
            "tone": "finalized",
        },
        {
            "label": "No Entregadas",
            "value": db.query(Requisicion).filter(
                or_(Requisicion.estado == "no_entregada", Requisicion.delivery_result == "no_entregada")
            ).count(),
            "tone": "rejected",
        },
    ]
    segmentos = []
    for segmento in segmentos_raw:
        porcentaje = round((segmento["value"] * 100 / total), 1) if total else 0
        segmentos.append(
            {
                **segmento,
                "percentage": porcentaje,
                "width_pct": max(porcentaje, 3) if segmento["value"] > 0 and total else 0,
            }
        )

    return {
        "total": total,
        "segments": segmentos,
        "has_data": total > 0,
    }


def build_home_aprobador_status_chart(current_user: Usuario, db: Session) -> dict[str, object] | None:
    if current_user.rol != "aprobador":
        return None

    pendientes_aprobacion = db.query(Requisicion).filter(Requisicion.estado == "pendiente").count()
    pendientes_entrega = db.query(Requisicion).filter(Requisicion.estado.in_(["aprobada", "preparado"])).count()
    pendientes_liquidacion = db.query(Requisicion).filter(
        Requisicion.estado == "entregada",
        or_(Requisicion.delivery_result.is_(None), Requisicion.delivery_result != "no_entregada"),
    ).count()
    finalizadas = db.query(Requisicion).filter(
        or_(
            Requisicion.estado.in_(["liquidada", "liquidada_en_prokey"]),
            Requisicion.estado == "no_entregada",
            Requisicion.delivery_result == "no_entregada",
        )
    ).count()
    rechazadas = db.query(Requisicion).filter(Requisicion.estado == "rechazada").count()
    total = pendientes_aprobacion + pendientes_entrega + pendientes_liquidacion + finalizadas + rechazadas
    segmentos_raw = [
        {
            "label": "Pendiente de aprobación",
            "value": pendientes_aprobacion,
            "tone": "pending",
        },
        {
            "label": "Pendiente de entrega",
            "value": pendientes_entrega,
            "tone": "process",
        },
        {
            "label": "Pendiente de liquidación",
            "value": pendientes_liquidacion,
            "tone": "closure",
        },
        {
            "label": "Finalizada",
            "value": finalizadas,
            "tone": "finalized",
        },
        {
            "label": "Rechazada",
            "value": rechazadas,
            "tone": "rejected",
        },
    ]
    segmentos = []
    for segmento in segmentos_raw:
        porcentaje = round((segmento["value"] * 100 / total), 1) if total else 0
        segmentos.append(
            {
                **segmento,
                "percentage": porcentaje,
                "width_pct": max(porcentaje, 3) if segmento["value"] > 0 and total else 0,
            }
        )

    return {
        "total": total,
        "segments": segmentos,
        "has_data": total > 0,
    }


def build_home_jefe_bodega_status_chart(current_user: Usuario, db: Session) -> dict[str, object] | None:
    if current_user.rol != "jefe_bodega":
        return None

    pendientes_aprobacion = db.query(Requisicion).filter(Requisicion.estado == "pendiente").count()
    pendientes_proceso = db.query(Requisicion).filter(Requisicion.estado.in_(["aprobada", "preparado"])).count()
    pendientes_liquidacion = db.query(Requisicion).filter(
        Requisicion.estado == "entregada",
        or_(Requisicion.delivery_result.is_(None), Requisicion.delivery_result != "no_entregada"),
    ).count()
    finalizadas = db.query(Requisicion).filter(
        or_(
            Requisicion.estado.in_(["liquidada", "liquidada_en_prokey"]),
            Requisicion.estado == "no_entregada",
            Requisicion.delivery_result == "no_entregada",
        )
    ).count()
    rechazadas = db.query(Requisicion).filter(Requisicion.estado == "rechazada").count()
    total = pendientes_aprobacion + pendientes_proceso + pendientes_liquidacion + finalizadas + rechazadas
    segmentos_raw = [
        {
            "label": "Pendiente de aprobación",
            "value": pendientes_aprobacion,
            "tone": "pending",
        },
        {
            "label": "Pendiente de proceso",
            "value": pendientes_proceso,
            "tone": "process",
        },
        {
            "label": "Pendiente de liquidación",
            "value": pendientes_liquidacion,
            "tone": "closure",
        },
        {
            "label": "Finalizada",
            "value": finalizadas,
            "tone": "finalized",
        },
        {
            "label": "Rechazada",
            "value": rechazadas,
            "tone": "rejected",
        },
    ]
    segmentos = []
    for segmento in segmentos_raw:
        porcentaje = round((segmento["value"] * 100 / total), 1) if total else 0
        segmentos.append(
            {
                **segmento,
                "percentage": porcentaje,
                "width_pct": max(porcentaje, 3) if segmento["value"] > 0 and total else 0,
            }
        )

    return {
        "total": total,
        "segments": segmentos,
        "has_data": total > 0,
    }


def build_home_jefe_bodega_monthly_chart(current_user: Usuario, db: Session) -> dict[str, object] | None:
    if current_user.rol != "jefe_bodega":
        return None

    ahora = now_sv()
    months = []
    year = ahora.year
    month = ahora.month
    for _ in range(6):
        months.append((year, month))
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    months.reverse()

    labels = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    bars = []
    max_value = 1
    for year_value, month_value in months:
        start = datetime(year_value, month_value, 1)
        if month_value == 12:
            end = datetime(year_value + 1, 1, 1)
        else:
            end = datetime(year_value, month_value + 1, 1)
        value = db.query(Requisicion).filter(Requisicion.created_at >= start, Requisicion.created_at < end).count()
        max_value = max(max_value, value)
        bars.append({"label": f"{labels[month_value - 1]} {str(year_value)[2:]}", "value": value})

    for bar in bars:
        bar["height_pct"] = 18 if bar["value"] == 0 else max(22, round((bar["value"] / max_value) * 100))

    return {
        "bars": bars,
        "has_data": any(bar["value"] > 0 for bar in bars),
        "max_value": max_value,
    }


def build_home_aprobador_monthly_chart(current_user: Usuario, db: Session) -> dict[str, object] | None:
    if current_user.rol != "aprobador":
        return None

    ahora = now_sv()
    months = []
    year = ahora.year
    month = ahora.month
    for _ in range(6):
        months.append((year, month))
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    months.reverse()

    labels = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    bars = []
    max_value = 1
    for year_value, month_value in months:
        start = datetime(year_value, month_value, 1)
        if month_value == 12:
            end = datetime(year_value + 1, 1, 1)
        else:
            end = datetime(year_value, month_value + 1, 1)
        count = (
            db.query(Requisicion)
            .filter(
                Requisicion.created_at >= start,
                Requisicion.created_at < end,
            )
            .count()
        )
        max_value = max(max_value, count)
        bars.append(
            {
                "label": f"{labels[month_value - 1]} {str(year_value)[-2:]}",
                "value": count,
            }
        )

    for bar in bars:
        bar["height_pct"] = max(round((bar["value"] * 100) / max_value, 1), 8) if bar["value"] > 0 else 0

    return {
        "bars": bars,
        "has_data": any(bar["value"] > 0 for bar in bars),
        "max_value": max_value,
    }


def build_home_aprobador_pending_age_chart(current_user: Usuario, db: Session) -> dict[str, object] | None:
    if current_user.rol != "aprobador":
        return None

    motivos = (
        db.query(
            Requisicion.motivo_requisicion,
            func.count(Requisicion.id).label("total"),
        )
        .filter(Requisicion.motivo_requisicion.is_not(None), Requisicion.motivo_requisicion != "")
        .group_by(Requisicion.motivo_requisicion)
        .order_by(func.count(Requisicion.id).desc(), Requisicion.motivo_requisicion.asc())
        .limit(5)
        .all()
    )

    buckets = []
    tones = ["fast", "medium", "slow", "very-slow", "process"]
    for index, row in enumerate(motivos):
        buckets.append(
            {
                "label": row[0],
                "value": row[1],
                "tone": tones[index % len(tones)],
            }
        )

    total = sum(bucket["value"] for bucket in buckets)
    max_value = max([1, *[bucket["value"] for bucket in buckets]])
    for bucket in buckets:
        bucket["height_pct"] = max(round((bucket["value"] * 100) / max_value, 1), 10) if bucket["value"] > 0 else 0
        bucket["percentage"] = round((bucket["value"] * 100) / total, 1) if total else 0

    return {
        "bars": buckets,
        "has_data": total > 0,
        "total": total,
    }


def build_home_bodega_monthly_chart(current_user: Usuario, db: Session) -> dict[str, object] | None:
    if current_user.rol != "bodega":
        return None

    ahora = now_sv()
    months = []
    year = ahora.year
    month = ahora.month
    for _ in range(6):
        months.append((year, month))
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    months.reverse()

    labels = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    bars = []
    max_value = 1
    for year_value, month_value in months:
        start = datetime(year_value, month_value, 1)
        if month_value == 12:
            end = datetime(year_value + 1, 1, 1)
        else:
            end = datetime(year_value, month_value + 1, 1)
        count = (
            db.query(Requisicion)
            .filter(
                Requisicion.delivered_at.is_not(None),
                Requisicion.delivered_at >= start,
                Requisicion.delivered_at < end,
            )
            .count()
        )
        max_value = max(max_value, count)
        bars.append(
            {
                "label": f"{labels[month_value - 1]} {str(year_value)[-2:]}",
                "value": count,
            }
        )

    for bar in bars:
        bar["height_pct"] = max(round((bar["value"] * 100) / max_value, 1), 8) if bar["value"] > 0 else 0

    return {
        "bars": bars,
        "has_data": any(bar["value"] > 0 for bar in bars),
        "max_value": max_value,
    }


def build_home_bodega_delivery_results_chart(current_user: Usuario, db: Session) -> dict[str, object] | None:
    if current_user.rol != "bodega":
        return None

    buckets = [
        {
            "label": "Completa",
            "value": db.query(Requisicion).filter(Requisicion.delivery_result == "completa").count(),
            "tone": "fast",
        },
        {
            "label": "Parcial",
            "value": db.query(Requisicion).filter(Requisicion.delivery_result == "parcial").count(),
            "tone": "slow",
        },
        {
            "label": "No Entregada",
            "value": db.query(Requisicion).filter(Requisicion.delivery_result == "no_entregada").count(),
            "tone": "very-slow",
        },
    ]

    total = sum(bucket["value"] for bucket in buckets)
    max_value = max(1, *(bucket["value"] for bucket in buckets))
    for bucket in buckets:
        bucket["height_pct"] = max(round((bucket["value"] * 100) / max_value, 1), 10) if bucket["value"] > 0 else 0
        bucket["percentage"] = round((bucket["value"] * 100) / total, 1) if total else 0

    return {
        "bars": buckets,
        "has_data": total > 0,
        "total": total,
    }


def build_home_jefe_bodega_delivery_results_chart(current_user: Usuario, db: Session) -> dict[str, object] | None:
    if current_user.rol != "jefe_bodega":
        return None

    buckets = [
        {
            "label": "Completa",
            "value": db.query(Requisicion).filter(Requisicion.delivery_result == "completa").count(),
            "tone": "fast",
        },
        {
            "label": "Parcial",
            "value": db.query(Requisicion).filter(Requisicion.delivery_result == "parcial").count(),
            "tone": "slow",
        },
        {
            "label": "No Entregada",
            "value": db.query(Requisicion).filter(Requisicion.delivery_result == "no_entregada").count(),
            "tone": "very-slow",
        },
    ]

    total = sum(bucket["value"] for bucket in buckets)
    max_value = max(1, *(bucket["value"] for bucket in buckets))
    for bucket in buckets:
        bucket["height_pct"] = max(round((bucket["value"] * 100) / max_value, 1), 10) if bucket["value"] > 0 else 0
        bucket["percentage"] = round((bucket["value"] * 100) / total, 1) if total else 0

    return {
        "bars": buckets,
        "has_data": total > 0,
        "total": total,
    }


def ensure_dashboard_access(current_user: Usuario) -> None:
    if current_user.rol not in ["admin", "aprobador", "jefe_bodega"]:
        raise HTTPException(status_code=403, detail="No autorizado")


def with_backup_operation_lock() -> bool:
    return BACKUP_OPERATION_LOCK.acquire(blocking=False)


def release_backup_operation_lock() -> None:
    if BACKUP_OPERATION_LOCK.locked():
        BACKUP_OPERATION_LOCK.release()


def build_users_import_report(rows: list[dict[str, str]], db: Session) -> dict[str, object]:
    report: dict[str, object] = {
        "total": len(rows),
        "created": 0,
        "updated": 0,
        "errors": 0,
        "error_rows": [],
        "preview_rows": [],
    }
    existing_users = db.query(Usuario).all()
    by_name = {normalize_person_name(u.nombre): u for u in existing_users}
    by_username = {u.username: u for u in existing_users}
    taken_usernames = set(by_username.keys())
    preview_rows: list[dict[str, str]] = []
    error_rows: list[dict[str, str]] = []

    for row in rows:
        nombre = normalize_text(row.get("nombre", ""))
        puesto = normalize_puesto(row.get("puesto", ""))
        linea = row.get("linea", "-")
        if len(nombre) < 3:
            error_rows.append({"linea": linea, "nombre": nombre, "puesto": puesto, "error": "Nombre invalido"})
            continue
        if puesto not in PUESTO_MAP:
            error_rows.append({"linea": linea, "nombre": nombre, "puesto": puesto, "error": "Puesto no mapeado"})
            continue

        rol, departamento = PUESTO_MAP[puesto]
        existing_by_name = by_name.get(normalize_person_name(nombre))
        created = existing_by_name is None
        if created:
            username_base = build_username_base(nombre)
            username = pick_unique_username(username_base, taken_usernames)
            by_username[username] = Usuario(
                username=username,
                nombre=nombre,
                rol=rol,
                departamento=departamento,
                activo=True,
                password="",
            )
        else:
            username = existing_by_name.username
        preview_rows.append(
            {
                "linea": linea,
                "nombre": nombre,
                "puesto": puesto,
                "rol": rol,
                "departamento": departamento,
                "username": username,
                "accion": "crear" if created else "actualizar",
            }
        )

    report["preview_rows"] = preview_rows
    report["error_rows"] = error_rows
    report["errors"] = len(error_rows)
    report["created"] = sum(1 for row in preview_rows if row["accion"] == "crear")
    report["updated"] = sum(1 for row in preview_rows if row["accion"] == "actualizar")
    return report


def apply_users_import(preview_rows: list[dict[str, str]], db: Session) -> dict[str, int]:
    created = 0
    updated = 0
    for row in preview_rows:
        username = row["username"]
        nombre = row["nombre"]
        rol = row["rol"]
        departamento = row["departamento"]
        usuario = db.query(Usuario).filter(Usuario.username == username).first()
        if usuario:
            usuario.nombre = nombre
            usuario.rol = rol
            usuario.departamento = departamento
            usuario.activo = True
            usuario.puede_iniciar_sesion = rol != "tecnico"
            if rol == "tecnico" and not usuario.pin_hash:
                usuario.pin_hash = hash_password(TEMP_IMPORT_PIN)
            updated += 1
            continue
        nuevo = Usuario(
            username=username,
            nombre=nombre,
            rol=rol,
            departamento=departamento,
            activo=True,
            password=hash_password(secrets.token_urlsafe(24)) if rol == "tecnico" else hash_password(TEMP_IMPORT_PASSWORD),
            pin_hash=hash_password(TEMP_IMPORT_PIN) if rol == "tecnico" else None,
            puede_iniciar_sesion=rol != "tecnico",
        )
        db.add(nuevo)
        created += 1
    db.commit()
    return {"created": created, "updated": updated}


def get_active_receptores(db: Session) -> list[Usuario]:
    return db.query(Usuario).filter(Usuario.activo.is_(True)).order_by(Usuario.nombre.asc()).all()


def validar_receptor_designado(db: Session, receptor_designado_id_raw: str) -> Usuario:
    receptor_designado_id_limpio = receptor_designado_id_raw.strip()
    if not receptor_designado_id_limpio:
        raise HTTPException(status_code=400, detail="Debes seleccionar receptor designado")
    try:
        receptor_designado_id_int = int(receptor_designado_id_limpio)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Receptor designado invalido") from exc
    receptor_designado = (
        db.query(Usuario)
        .filter(Usuario.id == receptor_designado_id_int, Usuario.activo.is_(True))
        .first()
    )
    if not receptor_designado:
        raise HTTPException(status_code=400, detail="El receptor designado no existe o esta inactivo")
    return receptor_designado


def safe_joinedload(model: type, attr_name: str):
    attr = getattr(model, attr_name, None)
    if attr is None:
        return None
    return joinedload(attr)


def validar_receptor_firma(
    db: Session,
    recibido_por_id_raw: str,
    pin_receptor: str,
    required: bool,
) -> tuple[Usuario | None, str | None]:
    recibido_por_id_limpio = str(recibido_por_id_raw or "").strip()
    pin_limpio = pin_receptor.strip()

    if not recibido_por_id_limpio and not pin_limpio:
        if required:
            return None, "Debes seleccionar receptor y escribir su PIN"
        return None, None

    if not recibido_por_id_limpio or not pin_limpio:
        return None, "Debes completar receptor y PIN"

    try:
        receptor_id = int(recibido_por_id_limpio)
    except ValueError:
        return None, "Receptor invalido"

    receptor = db.query(Usuario).filter(Usuario.id == receptor_id, Usuario.activo.is_(True)).first()
    if not receptor:
        return None, "El receptor seleccionado no existe o esta inactivo"
    if not receptor.pin_hash:
        return None, "El receptor seleccionado no tiene PIN configurado"
    if not verify_password(pin_limpio, receptor.pin_hash):
        return None, "PIN del receptor incorrecto"
    return receptor, None


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/login")
def login_form(request: Request):
    if request.session.get("user_id"):
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse("login.html", template_context(request))


@app.post("/login")
def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    username_clean = username.strip()
    user = authenticate_user(db, username=username, password=password)
    if not user:
        disabled_login_user = (
            db.query(Usuario)
            .filter(Usuario.username == username_clean, Usuario.activo.is_(True), Usuario.puede_iniciar_sesion.is_(False))
            .first()
        )
        logger.warning(
            "auth_login_failed",
            extra={
                "event": "auth_login_failed",
                "username": username_clean,
                "path": "/login",
                "method": "POST",
                "client_ip": get_client_ip(request),
                "reason": "sin_permiso_login" if disabled_login_user else "credenciales_invalidas",
            },
        )
        error_message = (
            "Este usuario no tiene permiso para iniciar sesion"
            if disabled_login_user
            else "Credenciales incorrectas"
        )
        return templates.TemplateResponse(
            "login.html",
            template_context(request, error=error_message),
            status_code=status.HTTP_401_UNAUTHORIZED,
        )
    login_user(request, user)
    logger.info(
        "auth_login_success",
        extra={
            "event": "auth_login_success",
            "user_id": user.id,
            "username": user.username,
            "role": user.rol,
            "path": "/login",
            "method": "POST",
            "client_ip": get_client_ip(request),
        },
    )
    return RedirectResponse(url="/", status_code=303)


@app.post("/logout")
def logout(request: Request):
    user_id = get_session_user_id(request)
    logout_user(request)
    logger.info(
        "auth_logout",
        extra={
            "event": "auth_logout",
            "user_id": user_id,
            "path": "/logout",
            "method": "POST",
            "client_ip": get_client_ip(request),
        },
    )
    return RedirectResponse(url="/login", status_code=303)


@app.get("/mi-cuenta/password")
def cambiar_password_form(
    request: Request,
    current_user: Usuario = Depends(get_current_user),
):
    if not current_user.puede_iniciar_sesion:
        raise HTTPException(status_code=403, detail="Este usuario no puede iniciar sesion")
    return templates.TemplateResponse(
        "cambiar_password.html",
        template_context(request, current_user),
    )


@app.post("/mi-cuenta/password")
def cambiar_password_guardar(
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not current_user.puede_iniciar_sesion:
        raise HTTPException(status_code=403, detail="Este usuario no puede iniciar sesion")

    current_clean = current_password.strip()
    new_clean = new_password.strip()
    confirm_clean = confirm_password.strip()

    if not verify_password(current_clean, current_user.password):
        return redirect_with_message("/mi-cuenta/password", "La contrasena actual no es correcta", "error")
    if len(new_clean) < PASSWORD_MIN_LENGTH:
        return redirect_with_message(
            "/mi-cuenta/password",
            f"La nueva contrasena debe tener al menos {PASSWORD_MIN_LENGTH} caracteres",
            "error",
        )
    if new_clean != confirm_clean:
        return redirect_with_message("/mi-cuenta/password", "La confirmacion no coincide", "error")
    if verify_password(new_clean, current_user.password):
        return redirect_with_message("/mi-cuenta/password", "La nueva contrasena debe ser diferente", "error")

    current_user.password = hash_password(new_clean)
    db.commit()
    return redirect_with_message("/mi-cuenta/password", "Contrasena actualizada", "success")


@app.get("/")
def home(request: Request, current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)):
    home_cards = build_home_cards(current_user, db)
    home_actions = build_home_actions(current_user)
    home_user_status_chart = build_home_user_status_chart(current_user, db)
    home_user_monthly_chart = build_home_user_monthly_chart(current_user, db)
    home_user_closure_chart = build_home_user_closure_chart(current_user, db)
    home_aprobador_status_chart = build_home_aprobador_status_chart(current_user, db)
    home_aprobador_monthly_chart = build_home_aprobador_monthly_chart(current_user, db)
    home_aprobador_pending_age_chart = build_home_aprobador_pending_age_chart(current_user, db)
    home_jefe_bodega_status_chart = build_home_jefe_bodega_status_chart(current_user, db)
    home_jefe_bodega_monthly_chart = build_home_jefe_bodega_monthly_chart(current_user, db)
    home_jefe_bodega_delivery_results_chart = build_home_jefe_bodega_delivery_results_chart(current_user, db)
    home_bodega_status_chart = build_home_bodega_status_chart(current_user, db)
    home_bodega_monthly_chart = build_home_bodega_monthly_chart(current_user, db)
    home_bodega_delivery_results_chart = build_home_bodega_delivery_results_chart(current_user, db)

    return templates.TemplateResponse(
        "home.html",
        template_context(
            request,
            current_user,
            home_cards=home_cards,
            home_actions=home_actions,
            home_user_status_chart=home_user_status_chart,
            home_user_monthly_chart=home_user_monthly_chart,
            home_user_closure_chart=home_user_closure_chart,
            home_aprobador_status_chart=home_aprobador_status_chart,
            home_aprobador_monthly_chart=home_aprobador_monthly_chart,
            home_aprobador_pending_age_chart=home_aprobador_pending_age_chart,
            home_jefe_bodega_status_chart=home_jefe_bodega_status_chart,
            home_jefe_bodega_monthly_chart=home_jefe_bodega_monthly_chart,
            home_jefe_bodega_delivery_results_chart=home_jefe_bodega_delivery_results_chart,
            home_bodega_status_chart=home_bodega_status_chart,
            home_bodega_monthly_chart=home_bodega_monthly_chart,
            home_bodega_delivery_results_chart=home_bodega_delivery_results_chart,
        ),
    )


@app.get("/monitor")
def get_monitor_actividad(request: Request, current_user: Usuario = Depends(get_current_user)):
    ensure_dashboard_access(current_user)
    return templates.TemplateResponse(
        "monitor_actividad.html",
        template_context(request, current_user),
    )


@app.get("/api/dashboard/basicos")
def dashboard_basicos_api(current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)):
    # Datos base para el Monitor de Actividad (Fase 1).
    ensure_dashboard_access(current_user)

    motivos_rows = (
        db.query(
            func.coalesce(Requisicion.motivo_requisicion, "Sin motivo").label("motivo"),
            func.count(Requisicion.id).label("total"),
        )
        .group_by(func.coalesce(Requisicion.motivo_requisicion, "Sin motivo"))
        .order_by(func.count(Requisicion.id).desc(), func.coalesce(Requisicion.motivo_requisicion, "Sin motivo").asc())
        .all()
    )

    solicitantes_rows = (
        db.query(
            Usuario.nombre.label("nombre"),
            func.count(Requisicion.id).label("total"),
        )
        .join(Usuario, Usuario.id == Requisicion.solicitante_id)
        .group_by(Usuario.id, Usuario.nombre)
        .order_by(func.count(Requisicion.id).desc(), Usuario.nombre.asc())
        .limit(10)
        .all()
    )

    items_rows = (
        db.query(
            Item.descripcion.label("descripcion"),
            func.sum(Item.cantidad).label("total_cantidad"),
        )
        .join(Requisicion, Requisicion.id == Item.requisicion_id)
        .group_by(Item.descripcion)
        .order_by(func.sum(Item.cantidad).desc(), Item.descripcion.asc())
        .limit(10)
        .all()
    )

    hourly_rows = (
        db.query(
            extract("hour", Requisicion.created_at).label("hora"),
            func.count(Requisicion.id).label("total"),
        )
        .group_by(extract("hour", Requisicion.created_at))
        .order_by(extract("hour", Requisicion.created_at).asc())
        .all()
    )
    heatmap_counts = {int(row.hora): int(row.total) for row in hourly_rows if row.hora is not None}
    heatmap_labels = [f"{hour:02d}:00" for hour in range(24)]
    heatmap_values = [heatmap_counts.get(hour, 0) for hour in range(24)]
    total_requisiciones = db.query(func.count(Requisicion.id)).scalar() or 0
    rango_fechas = db.query(
        func.min(Requisicion.created_at).label("min_created_at"),
        func.max(Requisicion.created_at).label("max_created_at"),
    ).one()
    dias_observados = 0
    if rango_fechas.min_created_at and rango_fechas.max_created_at:
        current_day = rango_fechas.min_created_at.date()
        end_day = rango_fechas.max_created_at.date()
        while current_day <= end_day:
            if current_day.weekday() < 5:
                dias_observados += 1
            current_day += timedelta(days=1)
    promedio_requisiciones_por_dia = round(total_requisiciones / dias_observados, 2) if dias_observados else 0.0

    prokey_cycle_rows = (
        db.query(Requisicion.created_at, Requisicion.prokey_liquidada_at)
        .filter(
            Requisicion.estado == "liquidada_en_prokey",
            Requisicion.created_at.is_not(None),
            Requisicion.prokey_liquidada_at.is_not(None),
        )
        .all()
    )
    cycle_hours = [
        max((row.prokey_liquidada_at - row.created_at).total_seconds() / 3600.0, 0.0)
        for row in prokey_cycle_rows
        if row.created_at and row.prokey_liquidada_at
    ]
    promedio_horas_hasta_prokey = round(sum(cycle_hours) / len(cycle_hours), 2) if cycle_hours else 0.0

    return {
        "kpis": {
            "promedio_horas_hasta_prokey": promedio_horas_hasta_prokey,
            "requisiciones_liquidadas_en_prokey": len(cycle_hours),
            "requisiciones_promedio_por_dia": promedio_requisiciones_por_dia,
            "dias_observados": dias_observados,
            "total_requisiciones": int(total_requisiciones),
        },
        "motivos": {
            "labels": [str(row.motivo or "Sin motivo") for row in motivos_rows],
            "values": [int(row.total or 0) for row in motivos_rows],
        },
        "top_solicitantes": {
            "labels": [str(row.nombre or "Sin nombre") for row in solicitantes_rows],
            "values": [int(row.total or 0) for row in solicitantes_rows],
        },
        "top_items": {
            "labels": [str(row.descripcion or "Sin descripcion") for row in items_rows],
            "values": [float(row.total_cantidad or 0) for row in items_rows],
        },
        "horario": {
            "labels": heatmap_labels,
            "values": heatmap_values,
            "alert_from_hour": 14,
        },
    }


@app.get("/api/dashboard/auditoria")
def dashboard_auditoria_api(current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)):
    # Datos de auditoria y diferencias para el Monitor de Actividad (Fase 2).
    ensure_dashboard_access(current_user)

    auditoria = build_dashboard_auditoria_snapshot(db)
    return {
        "kpis": auditoria["kpis"],
        "diferencia_por_producto": auditoria["diferencia_por_producto"],
        "diferencias_por_tecnico": auditoria["diferencias_por_tecnico"],
    }


def build_dashboard_auditoria_snapshot(db: Session) -> dict[str, object]:
    requisiciones_cerradas = (
        db.query(Requisicion)
        .options(
            joinedload(Requisicion.items),
            joinedload(Requisicion.solicitante),
            joinedload(Requisicion.receptor_designado),
        )
        .filter(Requisicion.estado.in_(["liquidada", "liquidada_en_prokey"]))
        .all()
    )

    total_requisiciones_cerradas = len(requisiciones_cerradas)
    requisiciones_con_diferencia = 0
    inversion_demos = 0.0
    diferencias_por_producto: dict[str, float] = {}
    diferencias_por_tecnico: dict[str, float] = {}
    drilldown_diferencias: list[dict[str, object]] = []
    drilldown_demos: list[dict[str, object]] = []

    for req in requisiciones_cerradas:
        req_tiene_diferencia = False
        total_diferencia_requisicion = 0.0
        items_con_diferencia = 0
        total_demo_requisicion = 0.0
        items_demo = 0
        receptor_nombre = (
            req.receptor_designado.nombre.strip()
            if req.receptor_designado and req.receptor_designado.nombre
            else "Sin receptor designado"
        )

        for item in req.items:
            delivered = float(item.cantidad_entregada or 0)
            used = float(item.qty_used or 0)
            not_used = float(item.qty_left_at_client or 0)
            returned = float(item.qty_returned_to_warehouse or 0)
            expected_return = float(
                calcular_retorno_esperado(
                    item.liquidation_mode,
                    used,
                    not_used,
                    item.contexto_operacion,
                )
            )
            diferencia = expected_return - returned

            if item.es_demo and delivered > 0:
                inversion_demos += delivered
                total_demo_requisicion += delivered
                items_demo += 1

            if diferencia <= 0:
                continue

            req_tiene_diferencia = True
            total_diferencia_requisicion += diferencia
            items_con_diferencia += 1
            descripcion = (item.descripcion or "Sin descripcion").strip() or "Sin descripcion"
            diferencias_por_producto[descripcion] = diferencias_por_producto.get(descripcion, 0.0) + diferencia
            diferencias_por_tecnico[receptor_nombre] = diferencias_por_tecnico.get(receptor_nombre, 0.0) + diferencia

        if req_tiene_diferencia:
            requisiciones_con_diferencia += 1
            drilldown_diferencias.append(
                {
                    "id": req.id,
                    "folio": req.folio,
                    "estado": req.estado,
                    "motivo_requisicion": req.motivo_requisicion,
                    "solicitante_nombre": req.solicitante.nombre if req.solicitante else "Sin solicitante",
                    "receptor_designado_nombre": receptor_nombre,
                    "liquidated_at": req.liquidated_at or req.prokey_liquidada_at or req.created_at,
                    "items_con_diferencia": items_con_diferencia,
                    "total_diferencia": round(total_diferencia_requisicion, 2),
                }
            )

        if items_demo > 0 and total_demo_requisicion > 0:
            drilldown_demos.append(
                {
                    "id": req.id,
                    "folio": req.folio,
                    "estado": req.estado,
                    "motivo_requisicion": req.motivo_requisicion,
                    "solicitante_nombre": req.solicitante.nombre if req.solicitante else "Sin solicitante",
                    "receptor_designado_nombre": receptor_nombre,
                    "liquidated_at": req.liquidated_at or req.prokey_liquidada_at or req.created_at,
                    "items_demo": items_demo,
                    "total_demo_entregado": round(total_demo_requisicion, 2),
                }
            )

    indice_discrepancia = (
        round((requisiciones_con_diferencia * 100.0) / total_requisiciones_cerradas, 2)
        if total_requisiciones_cerradas
        else 0.0
    )

    top_productos = sorted(
        diferencias_por_producto.items(),
        key=lambda item: (-item[1], item[0].lower()),
    )[:10]
    top_tecnicos = sorted(
        diferencias_por_tecnico.items(),
        key=lambda item: (-item[1], item[0].lower()),
    )[:5]

    drilldown_diferencias.sort(
        key=lambda item: (
            -float(item["total_diferencia"]),
            str(item["liquidated_at"] or ""),
            str(item["folio"]),
        )
    )
    drilldown_demos.sort(
        key=lambda item: (
            -float(item["total_demo_entregado"]),
            str(item["liquidated_at"] or ""),
            str(item["folio"]),
        )
    )

    return {
        "kpis": {
            "indice_discrepancia_pct": indice_discrepancia,
            "requisiciones_con_diferencia": requisiciones_con_diferencia,
            "requisiciones_cerradas": total_requisiciones_cerradas,
            "inversion_demos": round(inversion_demos, 2),
        },
        "diferencia_por_producto": {
            "labels": [label for label, _ in top_productos],
            "values": [round(value, 2) for _, value in top_productos],
        },
        "diferencias_por_tecnico": {
            "labels": [label for label, _ in top_tecnicos],
            "values": [round(value, 2) for _, value in top_tecnicos],
        },
        "requisiciones_con_diferencia_detalle": drilldown_diferencias,
        "requisiciones_demo_detalle": drilldown_demos,
    }


@app.get("/api/dashboard/auditoria/discrepancias")
def dashboard_auditoria_discrepancias_api(
    current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)
):
    ensure_dashboard_access(current_user)
    auditoria = build_dashboard_auditoria_snapshot(db)
    items = auditoria["requisiciones_con_diferencia_detalle"]
    return {
        "kind": "discrepancias",
        "title": "Requisiciones con diferencia",
        "description": "Requisiciones cerradas que presentan al menos un item con diferencia positiva.",
        "total": len(items),
        "items": items,
    }


@app.get("/api/dashboard/auditoria/demos")
def dashboard_auditoria_demos_api(current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_dashboard_access(current_user)
    auditoria = build_dashboard_auditoria_snapshot(db)
    items = auditoria["requisiciones_demo_detalle"]
    return {
        "kind": "demos",
        "title": "Requisiciones con demo",
        "description": "Requisiciones cerradas con al menos un item marcado para demo y cantidad entregada mayor a cero.",
        "total": len(items),
        "items": items,
    }


@app.get("/crear")
def crear_form(request: Request, current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)):
    restricted_redirect = redirect_if_bodega_plain_accesses_own_requests(current_user)
    if restricted_redirect:
        return restricted_redirect
    catalogo_items = (
        db.query(CatalogoItem)
        .filter(CatalogoItem.activo.is_(True))
        .order_by(CatalogoItem.nombre.asc())
        .all()
    )
    usuarios_activos = get_active_receptores(db)
    return templates.TemplateResponse(
        "crear_requisicion.html",
        template_context(
            request,
            current_user,
            catalogo_items=build_catalog_payload(catalogo_items),
            usuarios_activos=usuarios_activos,
            motivos_requisicion=MOTIVOS_REQUISICION,
        ),
    )


@app.post("/crear")
async def crear(
    request: Request,
    cliente_codigo: str = Form(...),
    cliente_nombre: str = Form(...),
    cliente_ruta_principal: str = Form(...),
    receptor_designado_id: str = Form(...),
    motivo_requisicion: str = Form(...),
    justificacion: str = Form(...),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    restricted_redirect = redirect_if_bodega_plain_accesses_own_requests(current_user)
    if restricted_redirect:
        return restricted_redirect
    cliente_codigo_limpio = cliente_codigo.strip()
    cliente_nombre_limpio = cliente_nombre.strip()
    cliente_ruta_principal_limpia = cliente_ruta_principal.strip().upper()
    if len(cliente_codigo_limpio) < 2:
        raise HTTPException(status_code=400, detail="Codigo de cliente invalido")
    if len(cliente_nombre_limpio) < 3:
        raise HTTPException(status_code=400, detail="Nombre de cliente invalido")
    if not re.fullmatch(r"[A-Z]{2}\d{2}", cliente_ruta_principal_limpia):
        raise HTTPException(status_code=400, detail="Ruta principal invalida (formato: AA00)")
    motivo_requisicion_limpio = motivo_requisicion.strip()
    if not motivo_requisicion_limpio:
        raise HTTPException(status_code=400, detail="Debes seleccionar un motivo")
    if motivo_requisicion_limpio not in MOTIVOS_REQUISICION:
        raise HTTPException(status_code=400, detail="Motivo de requisicion invalido")
    receptor_designado = validar_receptor_designado(db, receptor_designado_id)

    req = crear_requisicion_db(
        db=db,
        solicitante_id=current_user.id,
        departamento=current_user.departamento,
        cliente_codigo=cliente_codigo_limpio,
        cliente_nombre=cliente_nombre_limpio,
        cliente_ruta_principal=cliente_ruta_principal_limpia,
        motivo_requisicion=motivo_requisicion_limpio,
        justificacion=justificacion,
        receptor_designado_id=receptor_designado.id,
    )

    form_data = await request.form()
    try:
        items_data = parse_items_from_form(form_data)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not items_data:
        raise HTTPException(status_code=400, detail="Debe agregar al menos un item")
    descripciones = [normalize_catalog_name(item["descripcion"]) for item in items_data]
    if len(set(descripciones)) != len(descripciones):
        raise HTTPException(
            status_code=400,
            detail="No se permiten items duplicados en una misma requisicion",
        )

    catalogo_habilitado = {
        normalize_catalog_name(row.nombre): row
        for row in db.query(CatalogoItem).filter(CatalogoItem.activo.is_(True)).all()
    }
    if not catalogo_habilitado:
        raise HTTPException(status_code=400, detail="No hay items activos en catalogo")

    for item_data in items_data:
        descripcion_normalizada = normalize_catalog_name(item_data["descripcion"])
        if descripcion_normalizada not in catalogo_habilitado:
            raise HTTPException(status_code=400, detail="Item no permitido en catalogo")
        catalog_item = catalogo_habilitado[descripcion_normalizada]
        cantidad = float(item_data["cantidad"])
        if not catalog_item.permite_decimal and not cantidad.is_integer():
            raise HTTPException(
                status_code=400,
                detail=f"El item {catalog_item.nombre} solo permite cantidades enteras",
            )
        item_data["descripcion"] = catalog_item.nombre
        agregar_item_db(db, req.id, **item_data)

    return redirect_with_message("/mis-requisiciones", "Requisicion creada", "success")


@app.get("/mis-requisiciones")
def mis_requisiciones(
    request: Request, current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)
):
    restricted_redirect = redirect_if_bodega_plain_accesses_own_requests(current_user)
    if restricted_redirect:
        return restricted_redirect
    vista_param = request.query_params.get("vista", "mias").strip().lower()
    estado = request.query_params.get("estado", "todas").strip().lower()
    vista_global = puede_ver_todas_las_requisiciones(current_user) and vista_param == "todas"
    query = db.query(Requisicion).options(joinedload(Requisicion.solicitante))
    if not vista_global:
        query = query.filter(Requisicion.solicitante_id == current_user.id)

    estados_validos = {"pendiente", "aprobada", "preparado", "rechazada", "entregada", "no_entregada", "liquidada", "liquidada_en_prokey"}
    if estado == "abiertas":
        query = query.filter(
            Requisicion.estado.in_(["pendiente", "aprobada", "preparado", "entregada", "liquidada"]),
            or_(Requisicion.delivery_result.is_(None), Requisicion.delivery_result != "no_entregada"),
        )
    elif estado == "seguimiento":
        query = query.filter(
            Requisicion.estado.in_(["aprobada", "preparado", "entregada", "liquidada"]),
            or_(Requisicion.delivery_result.is_(None), Requisicion.delivery_result != "no_entregada"),
        )
    elif estado == "cerradas":
        query = query.filter(
            or_(
                Requisicion.estado == "liquidada_en_prokey",
                Requisicion.estado == "no_entregada",
                Requisicion.delivery_result == "no_entregada",
            )
        )
    elif estado in estados_validos:
        query = query.filter(Requisicion.estado == estado)

    requisiciones = (
        query
        .order_by(Requisicion.created_at.desc())
        .all()
    )
    return templates.TemplateResponse(
        "mis_requisiciones.html",
        template_context(
            request,
            current_user,
            requisiciones=requisiciones,
            vista_global=vista_global,
            filtro_estado=estado,
        ),
    )


@app.get("/mis-requisiciones/{req_id}/editar")
def editar_mi_requisicion_form(
    req_id: int,
    request: Request,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    restricted_redirect = redirect_if_bodega_plain_accesses_own_requests(current_user)
    if restricted_redirect:
        return restricted_redirect
    req = (
        db.query(Requisicion)
        .options(joinedload(Requisicion.items))
        .filter(Requisicion.id == req_id, Requisicion.solicitante_id == current_user.id)
        .first()
    )
    if not req:
        return redirect_with_message("/mis-requisiciones", "Requisicion no encontrada", "error")
    if req.estado != "pendiente" or req.approved_by is not None:
        return redirect_with_message(
            "/mis-requisiciones",
            "Solo puedes editar requisiciones pendientes sin aprobar",
            "warning",
        )

    catalogo_items = (
        db.query(CatalogoItem)
        .filter(CatalogoItem.activo.is_(True))
        .order_by(CatalogoItem.nombre.asc())
        .all()
    )
    return templates.TemplateResponse(
        "editar_requisicion.html",
        template_context(
            request,
            current_user,
            req=req,
            catalogo_items=build_catalog_payload(catalogo_items),
            usuarios_activos=get_active_receptores(db),
            motivos_requisicion=MOTIVOS_REQUISICION,
        ),
    )


@app.post("/mis-requisiciones/{req_id}/editar")
async def editar_mi_requisicion(
    req_id: int,
    request: Request,
    cliente_codigo: str = Form(...),
    cliente_nombre: str = Form(...),
    cliente_ruta_principal: str = Form(...),
    receptor_designado_id: str = Form(...),
    motivo_requisicion: str = Form(...),
    justificacion: str = Form(...),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    restricted_redirect = redirect_if_bodega_plain_accesses_own_requests(current_user)
    if restricted_redirect:
        return restricted_redirect
    req = (
        db.query(Requisicion)
        .options(joinedload(Requisicion.items))
        .filter(Requisicion.id == req_id, Requisicion.solicitante_id == current_user.id)
        .first()
    )
    if not req:
        return redirect_with_message("/mis-requisiciones", "Requisicion no encontrada", "error")
    if req.estado != "pendiente" or req.approved_by is not None:
        return redirect_with_message(
            "/mis-requisiciones",
            "Solo puedes editar requisiciones pendientes sin aprobar",
            "warning",
        )

    cliente_codigo_limpio = cliente_codigo.strip()
    cliente_nombre_limpio = cliente_nombre.strip()
    cliente_ruta_principal_limpia = cliente_ruta_principal.strip().upper()
    if len(cliente_codigo_limpio) < 2:
        raise HTTPException(status_code=400, detail="Codigo de cliente invalido")
    if len(cliente_nombre_limpio) < 3:
        raise HTTPException(status_code=400, detail="Nombre de cliente invalido")
    if not re.fullmatch(r"[A-Z]{2}\d{2}", cliente_ruta_principal_limpia):
        raise HTTPException(status_code=400, detail="Ruta principal invalida (formato: AA00)")

    motivo_requisicion_limpio = motivo_requisicion.strip()
    if not motivo_requisicion_limpio:
        raise HTTPException(status_code=400, detail="Debes seleccionar un motivo")
    if motivo_requisicion_limpio not in MOTIVOS_REQUISICION:
        raise HTTPException(status_code=400, detail="Motivo de requisicion invalido")

    receptor_designado = validar_receptor_designado(db, receptor_designado_id)

    form_data = await request.form()
    try:
        items_data = parse_items_from_form(form_data)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not items_data:
        raise HTTPException(status_code=400, detail="Debe agregar al menos un item")

    descripciones = [normalize_catalog_name(item["descripcion"]) for item in items_data]
    if len(set(descripciones)) != len(descripciones):
        raise HTTPException(
            status_code=400,
            detail="No se permiten items duplicados en una misma requisicion",
        )

    catalogo_habilitado = {
        normalize_catalog_name(row.nombre): row
        for row in db.query(CatalogoItem).filter(CatalogoItem.activo.is_(True)).all()
    }
    if not catalogo_habilitado:
        raise HTTPException(status_code=400, detail="No hay items activos en catalogo")

    for item_data in items_data:
        descripcion_normalizada = normalize_catalog_name(item_data["descripcion"])
        if descripcion_normalizada not in catalogo_habilitado:
            raise HTTPException(status_code=400, detail="Item no permitido en catalogo")
        catalog_item = catalogo_habilitado[descripcion_normalizada]
        cantidad = float(item_data["cantidad"])
        if not catalog_item.permite_decimal and not cantidad.is_integer():
            raise HTTPException(
                status_code=400,
                detail=f"El item {catalog_item.nombre} solo permite cantidades enteras",
            )
        item_data["descripcion"] = catalog_item.nombre

    req.cliente_codigo = cliente_codigo_limpio
    req.cliente_nombre = cliente_nombre_limpio
    req.cliente_ruta_principal = cliente_ruta_principal_limpia
    req.motivo_requisicion = motivo_requisicion_limpio
    req.justificacion = justificacion
    req.receptor_designado_id = receptor_designado.id

    req.items.clear()
    db.flush()
    for item_data in items_data:
        req.items.append(
            Item(
                descripcion=item_data["descripcion"],
                cantidad=float(item_data["cantidad"]),
                unidad=item_data["unidad"],
                contexto_operacion=item_data["contexto_operacion"],
                es_demo=bool(item_data["es_demo"]),
            )
        )

    db.commit()
    return redirect_with_message("/mis-requisiciones", "Requisicion actualizada", "success")


@app.post("/mis-requisiciones/{req_id}/eliminar")
def eliminar_mi_requisicion(
    req_id: int,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    restricted_redirect = redirect_if_bodega_plain_accesses_own_requests(current_user)
    if restricted_redirect:
        return restricted_redirect
    req = (
        db.query(Requisicion)
        .filter(Requisicion.id == req_id, Requisicion.solicitante_id == current_user.id)
        .first()
    )
    if not req:
        return redirect_with_message("/mis-requisiciones", "Requisicion no encontrada", "error")
    if req.estado != "pendiente":
        return redirect_with_message(
            "/mis-requisiciones",
            "Solo puedes eliminar requisiciones en pendiente de aprobar",
            "error",
        )

    db.delete(req)
    db.commit()
    return redirect_with_message("/mis-requisiciones", "Requisicion eliminada", "warning")


@app.get("/aprobar")
def aprobar_view(request: Request, current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.rol not in ["aprobador", "admin", "jefe_bodega"]:
        raise HTTPException(status_code=403, detail="No autorizado")

    q = request.query_params.get("q", "").strip()
    departamento = request.query_params.get("departamento", "todos").strip()
    query = (
        db.query(Requisicion)
        .options(
            joinedload(Requisicion.solicitante),
            joinedload(Requisicion.receptor_designado),
            joinedload(Requisicion.aprobador),
            joinedload(Requisicion.rechazador),
            joinedload(Requisicion.preparador),
            joinedload(Requisicion.entregador),
            joinedload(Requisicion.liquidator),
        )
        .filter(Requisicion.estado == "pendiente")
    )
    if departamento and departamento != "todos":
        query = query.filter(Requisicion.departamento == departamento)
    if q:
        patron = f"%{q}%"
        query = query.filter(
            or_(
                Requisicion.folio.ilike(patron),
                Requisicion.departamento.ilike(patron),
                Requisicion.justificacion.ilike(patron),
                Requisicion.cliente_codigo.ilike(patron),
                Requisicion.cliente_nombre.ilike(patron),
                Requisicion.solicitante.has(Usuario.nombre.ilike(patron)),
            )
        )

    requisiciones = query.order_by(Requisicion.created_at.desc()).all()
    departamentos = [
        row[0]
        for row in db.query(Requisicion.departamento).distinct().order_by(Requisicion.departamento.asc()).all()
        if row[0]
    ]
    return templates.TemplateResponse(
        "aprobar.html",
        template_context(
            request,
            current_user,
            requisiciones=requisiciones,
            filtro_q=q,
            filtro_departamento=departamento,
            departamentos=departamentos,
        ),
    )


@app.get("/todas-requisiciones")
def todas_requisiciones_view(
    request: Request, current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)
):
    ensure_all_requests_access(current_user)

    q = request.query_params.get("q", "").strip()
    estado = request.query_params.get("estado", "todos").strip().lower()
    departamento = request.query_params.get("departamento", "todos").strip()
    fecha_desde_raw = request.query_params.get("fecha_desde", "").strip()
    fecha_hasta_raw = request.query_params.get("fecha_hasta", "").strip()

    alias_estado = {
        "pendiente_aprobar": "pendiente",
        "pendiente_entregar": "aprobada",
    }
    estado_real = alias_estado.get(estado, estado)
    estados_validos = {"pendiente", "aprobada", "preparado", "rechazada", "entregada", "no_entregada", "liquidada", "liquidada_en_prokey"}
    query = (
        db.query(Requisicion)
        .options(
            joinedload(Requisicion.solicitante),
            joinedload(Requisicion.receptor_designado),
            joinedload(Requisicion.aprobador),
            joinedload(Requisicion.rechazador),
            joinedload(Requisicion.preparador),
            joinedload(Requisicion.entregador),
            joinedload(Requisicion.liquidator),
            joinedload(Requisicion.prokey_liquidator),
        )
        .filter(
            Requisicion.estado.in_(
                ["pendiente", "aprobada", "preparado", "rechazada", "entregada", "no_entregada", "liquidada", "liquidada_en_prokey"]
            )
        )
    )
    if estado == "pendiente_entregar":
        query = query.filter(Requisicion.estado.in_(["aprobada", "preparado"]))
    elif estado_real in estados_validos:
        query = query.filter(Requisicion.estado == estado_real)
    if departamento and departamento != "todos":
        query = query.filter(Requisicion.departamento == departamento)
    if fecha_desde_raw:
        try:
            fecha_desde = datetime.strptime(fecha_desde_raw, "%Y-%m-%d")
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Fecha desde invalida") from exc
        query = query.filter(Requisicion.created_at >= fecha_desde)
    if fecha_hasta_raw:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_raw, "%Y-%m-%d") + timedelta(days=1)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Fecha hasta invalida") from exc
        query = query.filter(Requisicion.created_at < fecha_hasta)
    if q:
        patron = f"%{q}%"
        query = query.filter(
            or_(
                Requisicion.folio.ilike(patron),
                Requisicion.departamento.ilike(patron),
                Requisicion.motivo_requisicion.ilike(patron),
                Requisicion.justificacion.ilike(patron),
                Requisicion.cliente_codigo.ilike(patron),
                Requisicion.cliente_nombre.ilike(patron),
                Requisicion.prokey_ref.ilike(patron),
                Requisicion.solicitante.has(Usuario.nombre.ilike(patron)),
                Requisicion.receptor_designado.has(Usuario.nombre.ilike(patron)),
                Requisicion.aprobador.has(Usuario.nombre.ilike(patron)),
                Requisicion.rechazador.has(Usuario.nombre.ilike(patron)),
                Requisicion.preparador.has(Usuario.nombre.ilike(patron)),
                Requisicion.entregador.has(Usuario.nombre.ilike(patron)),
                Requisicion.liquidator.has(Usuario.nombre.ilike(patron)),
                Requisicion.prokey_liquidator.has(Usuario.nombre.ilike(patron)),
            )
        )

    requisiciones = query.order_by(Requisicion.created_at.desc()).all()
    departamentos = [
        row[0]
        for row in db.query(Requisicion.departamento).distinct().order_by(Requisicion.departamento.asc()).all()
        if row[0]
    ]
    return templates.TemplateResponse(
        "todas_requisiciones.html",
        template_context(
            request,
            current_user,
            requisiciones=requisiciones,
            filtro_q=q,
            filtro_estado=estado,
            filtro_departamento=departamento,
            filtro_fecha_desde=fecha_desde_raw,
            filtro_fecha_hasta=fecha_hasta_raw,
            departamentos=departamentos,
        ),
    )


@app.get("/aprobar/{req_id}/gestionar")
def aprobar_gestionar(
    req_id: int,
    request: Request,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.rol not in ["aprobador", "admin", "jefe_bodega"]:
        raise HTTPException(status_code=403, detail="No autorizado")

    req = (
        db.query(Requisicion)
        .options(joinedload(Requisicion.solicitante), joinedload(Requisicion.items))
        .filter(Requisicion.id == req_id)
        .first()
    )
    if not req:
        raise HTTPException(status_code=404, detail="Requisicion no encontrada")
    if req.estado != "pendiente":
        return redirect_with_message("/aprobar", "Solo puedes gestionar requisiciones pendientes", "warning")
    if not puede_aprobar(req, current_user.rol):
        raise HTTPException(status_code=403, detail="No autorizado")

    return templates.TemplateResponse(
        "aprobar_gestionar.html",
        template_context(request, current_user, req=req),
    )


@app.post("/aprobar/{req_id}")
def aprobar(
    req_id: int,
    comentario: str = Form(""),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    req = db.query(Requisicion).filter(Requisicion.id == req_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="Requisicion no encontrada")
    if not puede_aprobar(req, current_user.rol):
        raise HTTPException(status_code=403, detail="No autorizado")

    transicionar_requisicion(
        db,
        req,
        nuevo_estado="aprobada",
        actor_id=current_user.id,
        approval_comment=comentario.strip() or None,
    )
    return redirect_with_message("/aprobar", "Requisicion aprobada", "success")


@app.post("/rechazar/{req_id}")
def rechazar(
    req_id: int,
    razon: str = Form(...),
    comentario: str = Form(""),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    req = db.query(Requisicion).filter(Requisicion.id == req_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="Requisicion no encontrada")
    if not puede_aprobar(req, current_user.rol):
        raise HTTPException(status_code=403, detail="No autorizado")

    razon_limpia = razon.strip()
    if len(razon_limpia) < 3:
        raise HTTPException(status_code=400, detail="La razon debe tener al menos 3 caracteres")

    transicionar_requisicion(
        db,
        req,
        nuevo_estado="rechazada",
        actor_id=current_user.id,
        rejection_reason=razon_limpia,
        rejection_comment=comentario.strip() or None,
    )
    return redirect_with_message("/aprobar", "Requisicion rechazada", "warning")


@app.get("/bodega")
def bodega_view(request: Request, current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.rol not in ["bodega", "admin", "jefe_bodega"]:
        raise HTTPException(status_code=403, detail="No autorizado")

    q = request.query_params.get("q", "").strip()
    vista = request.query_params.get("vista", "pendientes").strip().lower()
    if vista not in {"pendientes", "historial"}:
        vista = "pendientes"
    etapa = request.query_params.get("etapa", "todos").strip().lower()
    departamento = request.query_params.get("departamento", "todos").strip()
    resultado = request.query_params.get("resultado", "todos").strip().lower()
    fecha_desde_raw = request.query_params.get("fecha_desde", "").strip()
    fecha_hasta_raw = request.query_params.get("fecha_hasta", "").strip()

    bodega_optionals = [
        joinedload(Requisicion.solicitante),
        joinedload(Requisicion.receptor_designado),
        joinedload(Requisicion.aprobador),
        joinedload(Requisicion.preparador),
        joinedload(Requisicion.entregador),
        joinedload(Requisicion.liquidator),
    ]
    maybe_prokey_liquidator = safe_joinedload(Requisicion, "prokey_liquidator")
    if maybe_prokey_liquidator is not None:
        bodega_optionals.append(maybe_prokey_liquidator)

    pendientes_query = (
        db.query(Requisicion)
        .options(*bodega_optionals)
        .filter(
            or_(
                Requisicion.estado == "aprobada",
                Requisicion.estado == "preparado",
                Requisicion.estado == "entregada",
                Requisicion.estado == "liquidada",
            )
        )
    )
    pendientes_query = pendientes_query.filter(
        or_(
            Requisicion.estado == "aprobada",
            Requisicion.estado == "preparado",
            Requisicion.delivery_result.in_(["completa", "parcial"]),
            Requisicion.estado == "liquidada",
        )
    )
    if departamento and departamento != "todos":
        pendientes_query = pendientes_query.filter(Requisicion.departamento == departamento)
    if fecha_desde_raw:
        try:
            fecha_desde = datetime.strptime(fecha_desde_raw, "%Y-%m-%d")
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Fecha desde invalida") from exc
        pendientes_query = pendientes_query.filter(Requisicion.created_at >= fecha_desde)
    if fecha_hasta_raw:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_raw, "%Y-%m-%d") + timedelta(days=1)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Fecha hasta invalida") from exc
        pendientes_query = pendientes_query.filter(Requisicion.created_at < fecha_hasta)
    if etapa in {"aprobada", "preparado", "entregada", "liquidada"}:
        pendientes_query = pendientes_query.filter(Requisicion.estado == etapa)
    if q:
        patron = f"%{q}%"
        pendientes_query = pendientes_query.filter(
            or_(
                Requisicion.folio.ilike(patron),
                Requisicion.departamento.ilike(patron),
                Requisicion.justificacion.ilike(patron),
                Requisicion.cliente_codigo.ilike(patron),
                Requisicion.cliente_nombre.ilike(patron),
                Requisicion.solicitante.has(Usuario.nombre.ilike(patron)),
                Requisicion.receptor_designado.has(Usuario.nombre.ilike(patron)),
                Requisicion.aprobador.has(Usuario.nombre.ilike(patron)),
                Requisicion.preparador.has(Usuario.nombre.ilike(patron)),
                Requisicion.entregador.has(Usuario.nombre.ilike(patron)),
                Requisicion.liquidator.has(Usuario.nombre.ilike(patron)),
            )
        )

    pendientes_entrega = pendientes_query.order_by(
        Requisicion.approved_at.asc(),
        Requisicion.prepared_at.asc(),
        Requisicion.delivered_at.asc(),
        Requisicion.created_at.asc(),
    ).all()

    historial_query = (
        db.query(Requisicion)
        .options(
            joinedload(Requisicion.solicitante),
            joinedload(Requisicion.receptor_designado),
            joinedload(Requisicion.aprobador),
            joinedload(Requisicion.preparador),
            joinedload(Requisicion.entregador),
            joinedload(Requisicion.liquidator),
        )
        .filter(
            or_(
                Requisicion.estado == "liquidada_en_prokey",
                Requisicion.estado == "no_entregada",
                Requisicion.delivery_result == "no_entregada",
            )
        )
    )
    if departamento and departamento != "todos":
        historial_query = historial_query.filter(Requisicion.departamento == departamento)
    if fecha_desde_raw:
        try:
            fecha_desde = datetime.strptime(fecha_desde_raw, "%Y-%m-%d")
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Fecha desde invalida") from exc
        historial_query = historial_query.filter(Requisicion.created_at >= fecha_desde)
    if fecha_hasta_raw:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_raw, "%Y-%m-%d") + timedelta(days=1)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Fecha hasta invalida") from exc
        historial_query = historial_query.filter(Requisicion.created_at < fecha_hasta)
    if etapa == "liquidada_en_prokey":
        historial_query = historial_query.filter(Requisicion.estado == "liquidada_en_prokey")
    elif etapa == "no_entregada":
        historial_query = historial_query.filter(
            or_(Requisicion.estado == "no_entregada", Requisicion.delivery_result == "no_entregada")
        )
    if current_user.rol == "bodega":
        historial_query = historial_query.filter(
            or_(
                Requisicion.prepared_by == current_user.id,
                Requisicion.delivered_by == current_user.id,
                Requisicion.liquidated_by == current_user.id,
            )
        )
    # jefe_bodega ve el historial completo (sin filtro por usuario)
    if resultado in {"completa", "parcial", "no_entregada"}:
        historial_query = historial_query.filter(Requisicion.delivery_result == resultado)
    if q:
        patron = f"%{q}%"
        historial_query = historial_query.filter(
            or_(
                Requisicion.folio.ilike(patron),
                Requisicion.departamento.ilike(patron),
                Requisicion.justificacion.ilike(patron),
                Requisicion.cliente_codigo.ilike(patron),
                Requisicion.cliente_nombre.ilike(patron),
                Requisicion.solicitante.has(Usuario.nombre.ilike(patron)),
                Requisicion.receptor_designado.has(Usuario.nombre.ilike(patron)),
                Requisicion.aprobador.has(Usuario.nombre.ilike(patron)),
                Requisicion.preparador.has(Usuario.nombre.ilike(patron)),
                Requisicion.entregador.has(Usuario.nombre.ilike(patron)),
                Requisicion.liquidator.has(Usuario.nombre.ilike(patron)),
                Requisicion.delivered_to.ilike(patron),
            )
        )

    historial_entregadas = historial_query.order_by(Requisicion.delivered_at.desc()).all()
    departamentos = [
        row[0]
        for row in db.query(Requisicion.departamento).distinct().order_by(Requisicion.departamento.asc()).all()
        if row[0]
    ]

    return templates.TemplateResponse(
        "bodega.html",
        template_context(
            request,
            current_user,
            pendientes_entrega=pendientes_entrega,
            historial_entregadas=historial_entregadas,
            filtro_q=q,
            filtro_vista=vista,
            filtro_etapa=etapa,
            filtro_departamento=departamento,
            filtro_fecha_desde=fecha_desde_raw,
            filtro_fecha_hasta=fecha_hasta_raw,
            filtro_resultado=resultado,
            departamentos=departamentos,
        ),
    )


@app.get("/bodega/{req_id}/gestionar")
def bodega_gestionar(
    req_id: int,
    request: Request,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.rol not in ["bodega", "admin", "jefe_bodega"]:
        raise HTTPException(status_code=403, detail="No autorizado")

    req = (
        db.query(Requisicion)
        .options(
            joinedload(Requisicion.solicitante),
            joinedload(Requisicion.aprobador),
            joinedload(Requisicion.items),
            joinedload(Requisicion.receptor_designado),
        )
        .filter(Requisicion.id == req_id)
        .first()
    )
    if not req:
        raise HTTPException(status_code=404, detail="Requisicion no encontrada")
    if req.estado != "preparado":
        return redirect_with_message("/bodega", "Solo puedes gestionar requisiciones preparadas", "warning")
    if not puede_entregar(req, current_user.rol):
        raise HTTPException(status_code=403, detail="No autorizado")

    return templates.TemplateResponse(
        "bodega_gestionar.html",
        template_context(
            request,
            current_user,
            req=req,
            receptores=get_active_receptores(db),
            error_message=None,
            form_data={},
        ),
    )


@app.get("/bodega/{req_id}/preparar")
def bodega_preparar_form(
    req_id: int,
    request: Request,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    req = (
        db.query(Requisicion)
        .options(
            joinedload(Requisicion.items),
            joinedload(Requisicion.solicitante),
            joinedload(Requisicion.aprobador),
        )
        .filter(Requisicion.id == req_id)
        .first()
    )
    if not req:
        raise HTTPException(status_code=404, detail="Requisicion no encontrada")
    if not puede_preparar(req, current_user.rol):
        raise HTTPException(status_code=403, detail="No autorizado")

    return templates.TemplateResponse(
        "bodega_preparar.html",
        template_context(request, current_user, req=req),
    )


@app.post("/bodega/{req_id}/preparar")
def bodega_preparar_confirmar(
    req_id: int,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    req = (
        db.query(Requisicion)
        .options(joinedload(Requisicion.items))
        .filter(Requisicion.id == req_id)
        .first()
    )
    if not req:
        raise HTTPException(status_code=404, detail="Requisicion no encontrada")
    if not puede_preparar(req, current_user.rol):
        raise HTTPException(status_code=403, detail="No autorizado")

    transicionar_requisicion(
        db,
        req,
        nuevo_estado="preparado",
        actor_id=current_user.id,
    )
    return redirect_with_message("/bodega", "Requisicion marcada como preparada", "success")


@app.post("/requisiciones/{req_id}/liquidar-prokey")
def liquidar_en_prokey(
    req_id: int,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.rol not in ("jefe_bodega", "admin"):
        raise HTTPException(status_code=403, detail="No autorizado")
    try:
        marcar_liquidada_en_prokey(db, req_id=req_id, usuario_id=current_user.id)
    except ValueError as exc:
        return redirect_with_message("/bodega", str(exc), "error")
    return redirect_with_message("/bodega", "Requisicion confirmada como liquidada en Prokey", "success")


@app.post("/entregar/{req_id}")
def entregar(
    req_id: int,
    request: Request,
    resultado: str = Form("completa"),
    comentario: str = Form(""),
    recibido_por_id: str = Form(""),
    pin_receptor: str = Form(""),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    req = (
        db.query(Requisicion)
        .options(
            joinedload(Requisicion.items),
            joinedload(Requisicion.solicitante),
            joinedload(Requisicion.aprobador),
            joinedload(Requisicion.receptor_designado),
        )
        .filter(Requisicion.id == req_id)
        .first()
    )
    if not req:
        raise HTTPException(status_code=404, detail="Requisicion no encontrada")
    if not puede_entregar(req, current_user.rol):
        raise HTTPException(status_code=403, detail="No autorizado")

    resultados_validos = {"completa", "parcial", "no_entregada"}
    if resultado not in resultados_validos:
        return redirect_with_message("/bodega", "Resultado de entrega invalido", "error")

    if resultado == "parcial":
        return RedirectResponse(url=f"/entregar/{req_id}/parcial", status_code=303)

    comentario_limpio = comentario.strip()
    if resultado in {"parcial", "no_entregada"} and len(comentario_limpio) < 5:
        return redirect_with_message(
            "/bodega",
            "Para entrega parcial o no entregada, agrega comentario (minimo 5 caracteres)",
            "error",
        )

    requiere_firma = resultado in {"completa", "parcial"}
    receptor, firma_error = validar_receptor_firma(
        db,
        recibido_por_id if requiere_firma else "",
        pin_receptor if requiere_firma else "",
        required=requiere_firma,
    )
    if firma_error:
        return templates.TemplateResponse(
            "bodega_gestionar.html",
            template_context(
                request,
                current_user,
                req=req,
                receptores=get_active_receptores(db),
                error_message=firma_error,
                form_data={
                    "resultado": resultado,
                    "comentario": comentario,
                    "recibido_por_id": recibido_por_id,
                },
            ),
            status_code=400,
        )

    if resultado == "completa":
        for item in req.items:
            if item.cantidad_entregada is None:
                item.cantidad_entregada = item.cantidad

    nuevo_estado = "no_entregada" if resultado == "no_entregada" else "entregada"

    transicionar_requisicion(
        db,
        req,
        nuevo_estado=nuevo_estado,
        actor_id=current_user.id,
        delivered_to=receptor.nombre if receptor else None,
        delivery_result=resultado,
        delivery_comment=comentario_limpio or None,
        recibido_por_id=receptor.id if receptor else None,
        recibido_at=now_sv() if receptor else None,
    )
    if resultado == "no_entregada":
        for item in req.items:
            item.cantidad_entregada = 0
        db.commit()
    mensajes = {
        "completa": ("Requisicion marcada como entrega completa", "success"),
        "parcial": ("Requisicion marcada como entrega parcial", "warning"),
        "no_entregada": ("Requisicion registrada como no entregada", "warning"),
    }
    msg, tipo = mensajes[resultado]
    return redirect_with_message("/bodega", msg, tipo)


@app.get("/entregar/{req_id}/parcial")
def entrega_parcial_form(
    req_id: int,
    request: Request,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    req = (
        db.query(Requisicion)
        .options(
            joinedload(Requisicion.items),
            joinedload(Requisicion.solicitante),
            joinedload(Requisicion.aprobador),
            joinedload(Requisicion.receptor_designado),
        )
        .filter(Requisicion.id == req_id)
        .first()
    )
    if not req:
        raise HTTPException(status_code=404, detail="Requisicion no encontrada")
    if not puede_entregar(req, current_user.rol):
        raise HTTPException(status_code=403, detail="No autorizado")

    return templates.TemplateResponse(
        "bodega_entrega_parcial.html",
        template_context(
            request,
            current_user,
            req=req,
            receptores=get_active_receptores(db),
            error_message=None,
            form_data={},
            item_errors={},
        ),
    )


@app.post("/entregar/{req_id}/parcial")
async def entrega_parcial_guardar(
    req_id: int,
    request: Request,
    comentario: str = Form(...),
    recibido_por_id: str = Form(...),
    pin_receptor: str = Form(...),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    req = (
        db.query(Requisicion)
        .options(joinedload(Requisicion.items), joinedload(Requisicion.receptor_designado))
        .filter(Requisicion.id == req_id)
        .first()
    )
    if not req:
        raise HTTPException(status_code=404, detail="Requisicion no encontrada")
    if not puede_entregar(req, current_user.rol):
        raise HTTPException(status_code=403, detail="No autorizado")

    comentario_limpio = comentario.strip()
    if len(comentario_limpio) < 5:
        raise HTTPException(status_code=400, detail="Comentario minimo de 5 caracteres para entrega parcial")

    receptor, firma_error = validar_receptor_firma(db, recibido_por_id, pin_receptor, required=True)
    if firma_error:
        return templates.TemplateResponse(
            "bodega_entrega_parcial.html",
            template_context(
                request,
                current_user,
                req=req,
                receptores=get_active_receptores(db),
                error_message=firma_error,
                form_data={
                    "comentario": comentario,
                    "recibido_por_id": recibido_por_id,
                },
                item_errors={},
            ),
            status_code=400,
        )

    form_data = await request.form()
    total_entregado = 0.0
    parcial_detectada = False
    for item in req.items:
        raw = str(form_data.get(f"entregado_{item.id}", "0")).strip() or "0"
        try:
            entregado = float(raw)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Cantidad entregada invalida") from exc
        if entregado < 0:
            raise HTTPException(status_code=400, detail="Cantidad entregada no puede ser negativa")
        if entregado > item.cantidad:
            raise HTTPException(status_code=400, detail="Cantidad entregada no puede exceder la solicitada")
        item.cantidad_entregada = entregado
        total_entregado += entregado
        if entregado < item.cantidad:
            parcial_detectada = True

    if total_entregado <= 0:
        raise HTTPException(status_code=400, detail="Debe entregar al menos una cantidad para marcar parcial")
    if not parcial_detectada:
        raise HTTPException(status_code=400, detail="Si entregaste todo, usa resultado completa")

    transicionar_requisicion(
        db,
        req,
        nuevo_estado="entregada",
        actor_id=current_user.id,
        delivered_to=receptor.nombre,
        delivery_result="parcial",
        delivery_comment=comentario_limpio,
        recibido_por_id=receptor.id,
        recibido_at=now_sv(),
    )
    return redirect_with_message("/bodega", "Requisicion marcada como entrega parcial", "warning")


@app.get("/liquidar/{req_id}")
def liquidar_form(
    req_id: int,
    request: Request,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.rol not in ("admin", "bodega", "jefe_bodega"):
        raise HTTPException(status_code=403, detail="No autorizado")

    req = (
        db.query(Requisicion)
        .options(
            joinedload(Requisicion.items),
            joinedload(Requisicion.solicitante),
        )
        .filter(Requisicion.id == req_id)
        .first()
    )
    if not req:
        raise HTTPException(status_code=404, detail="Requisicion no encontrada")
    if req.estado == "liquidada":
        return redirect_with_message("/bodega", "Esta requisicion ya fue liquidada", "warning")
    if not puede_liquidar(req, current_user):
        return redirect_with_message("/bodega", "Requisicion no elegible para liquidacion", "error")

    attach_catalog_item_defaults(req.items, db)

    return templates.TemplateResponse(
        "liquidar.html",
        template_context(
            request,
            current_user,
            req=req,
            error_message=None,
            item_incompletos=[],
            liquidacion_values={},
            liquidacion_meta={"prokey_ref": "", "liquidation_comment": ""},
        ),
    )


@app.post("/liquidar/{req_id}")
async def liquidar_guardar(
    req_id: int,
    request: Request,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.rol not in ("admin", "bodega", "jefe_bodega"):
        raise HTTPException(status_code=403, detail="No autorizado")

    req = (
        db.query(Requisicion)
        .options(
            joinedload(Requisicion.items),
            joinedload(Requisicion.solicitante),
        )
        .filter(Requisicion.id == req_id)
        .first()
    )
    if not req:
        raise HTTPException(status_code=404, detail="Requisicion no encontrada")
    if req.estado == "liquidada":
        return redirect_with_message("/bodega", "Ya fue liquidada", "warning")
    if not puede_liquidar(req, current_user):
        return redirect_with_message("/bodega", "Requisicion no elegible para liquidacion", "error")

    attach_catalog_item_defaults(req.items, db)
    form_data = await request.form()
    prokey_ref = str(form_data.get("prokey_ref", "")).strip() or None
    liquidation_comment = str(form_data.get("liquidation_comment", "")).strip() or None
    liquidacion_meta = {
        "prokey_ref": str(form_data.get("prokey_ref", "")).strip(),
        "liquidation_comment": str(form_data.get("liquidation_comment", "")).strip(),
    }

    items_data: dict[int, dict[str, int | str | None]] = {}
    liquidacion_values: dict[int, dict[str, str]] = {}
    item_incompletos: list[int] = []
    item_validation_errors: dict[int, str] = {}

    def parse_non_negative_float(raw: str, field_label: str) -> float:
        try:
            value = round(float(raw), 4)
        except ValueError as exc:
            raise ValueError(f"{field_label} invalido") from exc
        if value < 0:
            raise ValueError(f"{field_label} no puede ser negativo")
        return value

    def is_effective_integer(value: float) -> bool:
        return float(value).is_integer()

    for item in req.items:
        qty_returned_raw = str(form_data.get(f"qty_returned_{item.id}", "0")).strip() or "0"
        qty_used_raw = str(form_data.get(f"qty_used_{item.id}", "0")).strip() or "0"
        qty_not_used_raw = str(form_data.get(f"qty_not_used_{item.id}", form_data.get(f"qty_left_{item.id}", "0"))).strip() or "0"
        locked_mode = str(getattr(item, "default_mode", "") or "").strip().upper()
        if locked_mode in ("RETORNABLE", "CONSUMIBLE"):
            mode_raw = locked_mode
        else:
            mode_raw = str(form_data.get(f"mode_{item.id}", "")).strip().upper()
        note_raw = str(form_data.get(f"note_{item.id}", "")).strip()
        note = note_raw or None
        liquidacion_values[item.id] = {
            "qty_returned": qty_returned_raw,
            "qty_used": qty_used_raw,
            "qty_not_used": qty_not_used_raw,
            "mode": mode_raw,
            "note": note_raw,
        }
        try:
            qty_returned = parse_non_negative_float(qty_returned_raw, "Regresa")
            qty_used = parse_non_negative_float(qty_used_raw, "Usado")
            qty_not_used = parse_non_negative_float(qty_not_used_raw, "No usado")
        except ValueError as exc:
            attach_catalog_item_defaults(req.items, db)
            return templates.TemplateResponse(
                "liquidar.html",
                template_context(
                    request,
                    current_user,
                    req=req,
                    error_message=str(exc),
                    item_incompletos=[],
                    liquidacion_values=liquidacion_values,
                    liquidacion_meta=liquidacion_meta,
                ),
                status_code=200,
            )
        if not bool(getattr(item, "permite_decimal", False)):
            if not all(is_effective_integer(value) for value in (qty_returned, qty_used, qty_not_used)):
                attach_catalog_item_defaults(req.items, db)
                return templates.TemplateResponse(
                    "liquidar.html",
                    template_context(
                        request,
                        current_user,
                        req=req,
                        error_message=f"El item {item.descripcion} solo permite cantidades enteras en liquidacion",
                        item_incompletos=[],
                        liquidacion_values=liquidacion_values,
                        liquidacion_meta=liquidacion_meta,
                    ),
                    status_code=200,
                )
        if mode_raw not in ("RETORNABLE", "CONSUMIBLE"):
            attach_catalog_item_defaults(req.items, db)
            return templates.TemplateResponse(
                "liquidar.html",
                template_context(
                    request,
                    current_user,
                    req=req,
                    error_message="Modo de liquidacion invalido",
                    item_incompletos=[],
                    liquidacion_values=liquidacion_values,
                    liquidacion_meta=liquidacion_meta,
                ),
                status_code=200,
            )

        delivered = round(float(item.cantidad_entregada or 0), 4)
        validation_error = validar_liquidacion_item(delivered, qty_used, qty_not_used, qty_returned, mode_raw)
        if validation_error:
            item_incompletos.append(item.id)
            item_validation_errors[item.id] = validation_error

        items_data[item.id] = {
            "qty_returned_to_warehouse": qty_returned,
            "qty_used": qty_used,
            "qty_left_at_client": qty_not_used,
            "liquidation_mode": mode_raw,
            "item_liquidation_note": note,
        }

    if item_incompletos:
        attach_catalog_item_defaults(req.items, db)
        return templates.TemplateResponse(
            "liquidar.html",
                template_context(
                    request,
                    current_user,
                    req=req,
                    error_message="Hay items con cobertura o retorno inconsistente. Revisa Usado, No usado y Regresa antes de liquidar.",
                    item_incompletos=item_incompletos,
                    item_validation_errors=item_validation_errors,
                    liquidacion_values=liquidacion_values,
                    liquidacion_meta=liquidacion_meta,
                ),
            status_code=200,
        )

    try:
        ejecutar_liquidacion(
            db=db,
            requisicion=req,
            usuario=current_user,
            prokey_ref=prokey_ref,
            liquidation_comment=liquidation_comment,
            items_data=items_data,
        )
    except ValueError as exc:
        return redirect_with_message("/bodega", str(exc), "warning")

    return redirect_with_message("/bodega", "Liquidacion registrada", "success")


@app.get("/requisiciones/{req_id}/prokey-ref")
def editar_prokey_ref_form(
    req_id: int,
    request: Request,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    req = (
        db.query(Requisicion)
        .options(joinedload(Requisicion.solicitante))
        .filter(Requisicion.id == req_id)
        .first()
    )
    if not req:
        raise HTTPException(status_code=404, detail="Requisicion no encontrada")

    if req.estado != "liquidada":
        return redirect_with_message("/mis-requisiciones", "Solo se puede completar referencia en requisiciones liquidadas", "error")
    if not puede_editar_prokey_ref(req, current_user):
        raise HTTPException(status_code=403, detail="No autorizado")

    cancel_url = "/bodega" if current_user.rol in ["admin", "jefe_bodega"] else "/mis-requisiciones"
    return templates.TemplateResponse(
        "editar_prokey_ref.html",
        template_context(request, current_user, req=req, cancel_url=cancel_url),
    )


@app.post("/requisiciones/{req_id}/prokey-ref")
async def editar_prokey_ref_guardar(
    req_id: int,
    request: Request,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    req = db.query(Requisicion).options(joinedload(Requisicion.items)).filter(Requisicion.id == req_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="Requisicion no encontrada")

    if req.estado != "liquidada":
        return redirect_with_message("/mis-requisiciones", "Solo se puede completar referencia en requisiciones liquidadas", "error")
    if not puede_editar_prokey_ref(req, current_user):
        raise HTTPException(status_code=403, detail="No autorizado")

    form_data = await request.form()
    prokey_ref = str(form_data.get("prokey_ref", "")).strip()
    if not prokey_ref:
        return redirect_with_message(f"/requisiciones/{req_id}/prokey-ref", "La referencia Prokey es obligatoria", "error")

    req.prokey_ref = prokey_ref
    req.prokey_ref_actualizada_at = now_sv()
    req.prokey_ref_actualizada_por = current_user.id
    db.commit()

    target = "/bodega" if current_user.rol in ["admin", "jefe_bodega"] else "/mis-requisiciones"
    return redirect_with_message(target, "Referencia Prokey actualizada", "success")


@app.get("/admin/usuarios")
def admin_usuarios(request: Request, current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_admin(current_user)
    estado = request.query_params.get("estado", "activos")
    query = db.query(Usuario)
    if estado == "activos":
        query = query.filter(Usuario.activo.is_(True))
    elif estado == "inactivos":
        query = query.filter(Usuario.activo.is_(False))
    usuarios = query.order_by(Usuario.activo.desc(), Usuario.id.asc()).all()
    return templates.TemplateResponse(
        "admin_usuarios.html",
        template_context(request, current_user, usuarios=usuarios, estado=estado, import_report=None),
    )


@app.get("/admin/respaldos")
def admin_respaldos(request: Request, current_user: Usuario = Depends(get_current_user)):
    ensure_admin(current_user)
    return templates.TemplateResponse(
        "admin_respaldos.html",
        template_context(
            request,
            current_user,
            backups_supported=is_sqlite_database(),
            backups=build_backup_rows() if is_sqlite_database() else [],
            backup_dir=str(get_backup_directory()),
        ),
    )


@app.post("/admin/respaldos/generar")
def admin_respaldos_generar(request: Request, current_user: Usuario = Depends(get_current_user)):
    ensure_admin(current_user)
    if not is_sqlite_database():
        return redirect_with_message("/admin/respaldos", "Los respaldos automaticos solo soportan SQLite", "error")
    if not with_backup_operation_lock():
        return redirect_with_message("/admin/respaldos", "Ya hay una operacion de respaldo o restauracion en curso", "warning")
    try:
        archive_path = create_backup_archive()
        logger.info(
            "admin_backup_created",
            extra={
                "event": "admin_backup_created",
                "user_id": current_user.id,
                "backup_filename": archive_path.name,
            },
        )
        return FileResponse(
            path=str(archive_path),
            media_type="application/zip",
            filename=archive_path.name,
        )
    except Exception as exc:
        logger.exception(
            "admin_backup_failed",
            extra={"event": "admin_backup_failed", "user_id": current_user.id},
        )
        return redirect_with_message("/admin/respaldos", f"No se pudo generar el respaldo: {exc}", "error")
    finally:
        release_backup_operation_lock()


@app.get("/admin/respaldos/{backup_name}/descargar")
def admin_respaldos_descargar(
    backup_name: str,
    current_user: Usuario = Depends(get_current_user),
):
    ensure_admin(current_user)
    try:
        archive_path = resolve_backup_archive(backup_name)
    except Exception as exc:
        return redirect_with_message("/admin/respaldos", str(exc), "error")
    return FileResponse(path=str(archive_path), media_type="application/zip", filename=archive_path.name)


def _perform_restore(
    *,
    request: Request,
    current_user: Usuario,
    archive_path,
    source_label: str,
) -> RedirectResponse:
    if not with_backup_operation_lock():
        return redirect_with_message("/admin/respaldos", "Ya hay una operacion de respaldo o restauracion en curso", "warning")

    MAINTENANCE_STATE["active"] = True
    MAINTENANCE_STATE["reason"] = "La aplicacion esta restaurando un respaldo administrativo. Vuelve a ingresar en unos segundos."
    try:
        result = restore_backup_archive(archive_path)
        logger.warning(
            "admin_restore_completed",
            extra={
                "event": "admin_restore_completed",
                "user_id": current_user.id,
                "backup_source": source_label,
                "safety_backup": result["safety_backup"],
            },
        )
    except Exception as exc:
        logger.exception(
            "admin_restore_failed",
            extra={
                "event": "admin_restore_failed",
                "user_id": current_user.id,
                "backup_source": source_label,
            },
        )
        return redirect_with_message("/admin/respaldos", f"No se pudo restaurar el respaldo: {exc}", "error")
    finally:
        MAINTENANCE_STATE["active"] = False
        MAINTENANCE_STATE["reason"] = None
        release_backup_operation_lock()

    request.session.clear()
    return redirect_with_message(
        "/login",
        f"Respaldo restaurado. Backup de seguridad previo: {result['safety_backup']}. Inicia sesion nuevamente.",
        "success",
    )


@app.post("/admin/respaldos/{backup_name}/restaurar")
def admin_respaldos_restaurar_guardado(
    backup_name: str,
    request: Request,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    if not is_sqlite_database():
        return redirect_with_message("/admin/respaldos", "La restauracion automatica solo soporta SQLite", "error")
    try:
        archive_path = resolve_backup_archive(backup_name)
    except Exception as exc:
        return redirect_with_message("/admin/respaldos", str(exc), "error")
    db.close()
    return _perform_restore(
        request=request,
        current_user=current_user,
        archive_path=archive_path,
        source_label=archive_path.name,
    )


@app.post("/admin/respaldos/restaurar")
async def admin_respaldos_restaurar_subido(
    request: Request,
    file: UploadFile = File(...),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    if not is_sqlite_database():
        return redirect_with_message("/admin/respaldos", "La restauracion automatica solo soporta SQLite", "error")
    if not file.filename:
        return redirect_with_message("/admin/respaldos", "Debes seleccionar un archivo ZIP", "error")
    if not file.filename.lower().endswith(".zip"):
        return redirect_with_message("/admin/respaldos", "Solo se aceptan archivos ZIP de respaldo", "error")
    raw = await file.read()
    if not raw:
        return redirect_with_message("/admin/respaldos", "El archivo de respaldo esta vacio", "error")
    db.close()
    with tempfile.NamedTemporaryFile(prefix="restore-upload-", suffix=".zip", delete=False) as temp_file:
        temp_file.write(raw)
        temp_path = temp_file.name
    try:
        return _perform_restore(
            request=request,
            current_user=current_user,
            archive_path=temp_path,
            source_label=file.filename,
        )
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


@app.post("/admin/usuarios/importar")
async def admin_usuarios_importar(
    request: Request,
    file: UploadFile = File(...),
    dry_run: str = Form("1"),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    if not file.filename:
        return redirect_with_message("/admin/usuarios", "Debes seleccionar un archivo", "error")
    raw = await file.read()
    if not raw:
        return redirect_with_message("/admin/usuarios", "El archivo esta vacio", "error")
    try:
        rows = _parse_users_rows(raw, file.filename)
        report = build_users_import_report(rows, db)
    except ValueError as exc:
        return redirect_with_message("/admin/usuarios", str(exc), "error")

    usuarios = db.query(Usuario).filter(Usuario.activo.is_(True)).order_by(Usuario.activo.desc(), Usuario.id.asc()).all()
    if dry_run == "1":
        return templates.TemplateResponse(
            "admin_usuarios.html",
            template_context(request, current_user, usuarios=usuarios, estado="activos", import_report=report),
        )

    if report["errors"]:
        return templates.TemplateResponse(
            "admin_usuarios.html",
            template_context(request, current_user, usuarios=usuarios, estado="activos", import_report=report),
        )

    result = apply_users_import(report["preview_rows"], db)
    return redirect_with_message(
        "/admin/usuarios",
        f"Importacion completada: {result['created']} creados, {result['updated']} actualizados",
        "success",
    )


@app.get("/admin/usuarios/nuevo")
def admin_usuario_nuevo(request: Request, current_user: Usuario = Depends(get_current_user)):
    ensure_admin(current_user)
    return templates.TemplateResponse(
        "admin_usuario_form.html",
        template_context(
            request,
            current_user,
            modo="crear",
            usuario=None,
            roles=ROLES_VALIDOS,
            departamentos=DEPARTAMENTOS_VALIDOS,
        ),
    )


@app.post("/admin/usuarios")
def admin_usuario_crear(
    request: Request,
    username: str = Form(...),
    nombre: str = Form(...),
    rol: str = Form(...),
    departamento: str = Form(...),
    password: str = Form(""),
    pin: str = Form(""),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    username_limpio = username.strip()
    nombre_limpio = nombre.strip()
    depto_limpio = departamento.strip()
    pin_limpio = pin.strip()

    if rol not in ROLES_VALIDOS:
        raise HTTPException(status_code=400, detail="Rol invalido")
    if depto_limpio not in DEPARTAMENTOS_VALIDOS:
        raise HTTPException(status_code=400, detail="Departamento invalido")
    if len(username_limpio) < 3 or len(nombre_limpio) < 3 or len(depto_limpio) < 2:
        raise HTTPException(status_code=400, detail="Datos incompletos o invalidos")
    if rol == "tecnico":
        if not pin_limpio:
            raise HTTPException(status_code=400, detail="El PIN es obligatorio para usuarios tecnicos")
    elif len(password.strip()) < 6:
        raise HTTPException(status_code=400, detail="La contrasena debe tener al menos 6 caracteres")

    existe = db.query(Usuario).filter(Usuario.username == username_limpio).first()
    if existe:
        return redirect_with_message("/admin/usuarios", "Username ya existe", "error")

    nuevo = Usuario(
        username=username_limpio,
        nombre=nombre_limpio,
        rol=rol,
        departamento=depto_limpio,
        activo=True,
        password=hash_password(password.strip()) if password.strip() else hash_password(secrets.token_urlsafe(24)),
        pin_hash=hash_password(pin_limpio) if pin_limpio else None,
        puede_iniciar_sesion=rol != "tecnico",
    )
    db.add(nuevo)
    db.commit()
    return redirect_with_message("/admin/usuarios", "Usuario creado", "success")


@app.get("/admin/usuarios/{user_id}/editar")
def admin_usuario_editar_form(
    user_id: int,
    request: Request,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    usuario = db.query(Usuario).filter(Usuario.id == user_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return templates.TemplateResponse(
        "admin_usuario_form.html",
        template_context(
            request,
            current_user,
            modo="editar",
            usuario=usuario,
            roles=ROLES_VALIDOS,
            departamentos=DEPARTAMENTOS_VALIDOS,
        ),
    )


@app.post("/admin/usuarios/{user_id}/editar")
def admin_usuario_editar(
    user_id: int,
    username: str = Form(...),
    nombre: str = Form(...),
    rol: str = Form(...),
    departamento: str = Form(...),
    password: str = Form(""),
    pin: str = Form(""),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    usuario = db.query(Usuario).filter(Usuario.id == user_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    username_limpio = username.strip()
    nombre_limpio = nombre.strip()
    depto_limpio = departamento.strip()
    pin_limpio = pin.strip()
    if rol not in ROLES_VALIDOS:
        raise HTTPException(status_code=400, detail="Rol invalido")
    if depto_limpio not in DEPARTAMENTOS_VALIDOS:
        raise HTTPException(status_code=400, detail="Departamento invalido")
    if len(username_limpio) < 3 or len(nombre_limpio) < 3 or len(depto_limpio) < 2:
        raise HTTPException(status_code=400, detail="Datos incompletos o invalidos")
    if rol == "tecnico" and not (pin_limpio or usuario.pin_hash):
        raise HTTPException(status_code=400, detail="El PIN es obligatorio para usuarios tecnicos")

    duplicado = db.query(Usuario).filter(Usuario.username == username_limpio, Usuario.id != user_id).first()
    if duplicado:
        return redirect_with_message("/admin/usuarios", "Username ya existe", "error")

    usuario.username = username_limpio
    usuario.nombre = nombre_limpio
    usuario.rol = rol
    usuario.departamento = depto_limpio
    usuario.puede_iniciar_sesion = rol != "tecnico"

    if password.strip():
        if len(password.strip()) < 6:
            raise HTTPException(status_code=400, detail="La contrasena debe tener al menos 6 caracteres")
        usuario.password = hash_password(password.strip())
    if pin_limpio:
        usuario.pin_hash = hash_password(pin_limpio)

    db.commit()
    return redirect_with_message("/admin/usuarios", "Usuario actualizado", "success")


@app.post("/admin/requisiciones/{req_id}/eliminar")
def admin_requisicion_eliminar(
    req_id: int,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    req = db.query(Requisicion).filter(Requisicion.id == req_id).first()
    if not req:
        return redirect_with_message("/aprobar", "Requisicion no encontrada", "error")
    folio = req.folio
    db.delete(req)
    db.commit()
    return redirect_with_message("/aprobar", f"Requisicion {folio} eliminada", "warning")


@app.post("/admin/usuarios/{user_id}/eliminar")
def admin_usuario_eliminar(
    user_id: int,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    usuario = db.query(Usuario).filter(Usuario.id == user_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if usuario.id == current_user.id:
        return redirect_with_message("/admin/usuarios", "No puedes eliminar tu propio usuario", "error")

    if usuario.rol == "admin":
        total_admins = db.query(Usuario).filter(Usuario.rol == "admin").count()
        if total_admins <= 1:
            return redirect_with_message("/admin/usuarios", "No puedes eliminar el ultimo admin", "error")

    referencias = (
        db.query(Requisicion.id)
        .filter(
            or_(
                Requisicion.solicitante_id == usuario.id,
                Requisicion.approved_by == usuario.id,
                Requisicion.rejected_by == usuario.id,
                Requisicion.delivered_by == usuario.id,
            )
        )
        .first()
    )
    if referencias:
        return redirect_with_message(
            "/admin/usuarios",
            "No puedes eliminar un usuario con historial en requisiciones",
            "error",
        )

    db.delete(usuario)
    db.commit()
    return redirect_with_message("/admin/usuarios", "Usuario eliminado", "warning")


@app.post("/admin/usuarios/{user_id}/desactivar")
def admin_usuario_desactivar(
    user_id: int,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    usuario = db.query(Usuario).filter(Usuario.id == user_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if usuario.id == current_user.id:
        return redirect_with_message("/admin/usuarios", "No puedes desactivar tu propio usuario", "error")
    if not usuario.activo:
        return redirect_with_message("/admin/usuarios", "Usuario ya inactivo", "warning")

    usuario.activo = False
    db.commit()
    return redirect_with_message("/admin/usuarios", "Usuario desactivado", "warning")


@app.post("/admin/usuarios/{user_id}/reactivar")
def admin_usuario_reactivar(
    user_id: int,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    usuario = db.query(Usuario).filter(Usuario.id == user_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if usuario.activo:
        return redirect_with_message("/admin/usuarios", "Usuario ya activo", "warning")

    usuario.activo = True
    db.commit()
    return redirect_with_message("/admin/usuarios", "Usuario reactivado", "success")


@app.get("/admin/catalogo-items")
def admin_catalogo_items(
    request: Request,
    q: str = "",
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    q_limpio = q.strip()
    query = db.query(CatalogoItem)
    if q_limpio:
        q_norm = q_limpio.lower()
        query = query.filter(func.lower(CatalogoItem.nombre).like(f"%{q_norm}%"))
    items = query.order_by(CatalogoItem.nombre.asc()).all()
    return templates.TemplateResponse(
        "admin_catalogo_items.html",
        template_context(request, current_user, items=items, q=q_limpio),
    )


@app.post("/admin/catalogo-items/eliminar-todos")
def admin_catalogo_items_eliminar_todos(
    confirmacion_texto: str = Form(""),
    confirmacion_check: str | None = Form(None),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)

    if confirmacion_check != "on":
        return redirect_with_message(
            "/admin/catalogo-items",
            "Debes confirmar explicitamente que deseas borrar todo el catalogo",
            "error",
        )

    if confirmacion_texto.strip().upper() != "BORRAR CATALOGO":
        return redirect_with_message(
            "/admin/catalogo-items",
            "Escribe exactamente BORRAR CATALOGO para confirmar",
            "error",
        )

    total = db.query(CatalogoItem).count()
    if total == 0:
        return redirect_with_message("/admin/catalogo-items", "El catalogo ya esta vacio", "warning")

    db.query(CatalogoItem).delete()
    db.commit()
    return redirect_with_message(
        "/admin/catalogo-items",
        f"Catalogo eliminado por completo ({total} items)",
        "warning",
    )


@app.post("/admin/catalogo-items/importar")
async def admin_catalogo_item_importar(
    archivo: UploadFile = File(...),
    activar_importados: str | None = Form(None),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    filename = (archivo.filename or "").strip()
    if not filename:
        return redirect_with_message("/admin/catalogo-items", "Debes seleccionar un archivo", "error")

    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext not in {"csv", "xlsx"}:
        return redirect_with_message("/admin/catalogo-items", "Formato no soportado. Usa CSV o XLSX", "error")

    raw = await archivo.read()
    if not raw:
        return redirect_with_message("/admin/catalogo-items", "El archivo esta vacio", "error")

    try:
        if ext == "csv":
            nombres = _parse_catalog_csv(raw)
        else:
            nombres = _parse_catalog_xlsx(raw)
    except ValueError as exc:
        return redirect_with_message("/admin/catalogo-items", str(exc), "error")

    if not nombres:
        return redirect_with_message("/admin/catalogo-items", "No se encontraron nombres de items validos", "error")

    activar = activar_importados == "on"
    existentes = {normalize_catalog_name(i.nombre): i for i in db.query(CatalogoItem).all()}
    creados = 0
    actualizados = 0
    sin_cambios = 0

    for nombre in nombres:
        key = normalize_catalog_name(nombre)
        existente = existentes.get(key)
        tipo_item = classify_catalog_item_type(nombre)
        permite_decimal = catalog_item_allows_decimal(nombre)
        if existente:
            hubo_cambio = False
            if existente.activo != activar:
                existente.activo = activar
                hubo_cambio = True
            if existente.tipo_item != tipo_item:
                existente.tipo_item = tipo_item
                hubo_cambio = True
            if existente.permite_decimal != permite_decimal:
                existente.permite_decimal = permite_decimal
                hubo_cambio = True
            if hubo_cambio:
                actualizados += 1
            else:
                sin_cambios += 1
            continue
        db.add(
            CatalogoItem(
                nombre=nombre,
                activo=activar,
                tipo_item=tipo_item,
                permite_decimal=permite_decimal,
            )
        )
        creados += 1

    db.commit()
    msg = f"Importacion completada. Creados: {creados}, actualizados: {actualizados}, sin cambios: {sin_cambios}"
    return redirect_with_message("/admin/catalogo-items", msg, "success")


@app.get("/admin/catalogo-items/nuevo")
def admin_catalogo_item_nuevo(request: Request, current_user: Usuario = Depends(get_current_user)):
    ensure_admin(current_user)
    return templates.TemplateResponse(
        "admin_catalogo_item_form.html",
        template_context(request, current_user, modo="crear", item=None),
    )


@app.post("/admin/catalogo-items")
def admin_catalogo_item_crear(
    nombre: str = Form(...),
    activo: str = Form("on"),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    nombre_limpio = nombre.strip()
    if len(nombre_limpio) < 2:
        raise HTTPException(status_code=400, detail="Nombre de item invalido")

    existe = db.query(CatalogoItem).filter(CatalogoItem.nombre == nombre_limpio).first()
    if existe:
        return redirect_with_message("/admin/catalogo-items", "El item ya existe", "error")

    nuevo = CatalogoItem(
        nombre=nombre_limpio,
        activo=(activo == "on"),
        tipo_item=classify_catalog_item_type(nombre_limpio),
        permite_decimal=catalog_item_allows_decimal(nombre_limpio),
    )
    db.add(nuevo)
    db.commit()
    return redirect_with_message("/admin/catalogo-items", "Item creado", "success")


@app.get("/admin/catalogo-items/{item_id}/editar")
def admin_catalogo_item_editar_form(
    item_id: int,
    request: Request,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    item = db.query(CatalogoItem).filter(CatalogoItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item no encontrado")
    return templates.TemplateResponse(
        "admin_catalogo_item_form.html",
        template_context(request, current_user, modo="editar", item=item),
    )


@app.post("/admin/catalogo-items/{item_id}/editar")
def admin_catalogo_item_editar(
    item_id: int,
    nombre: str = Form(...),
    activo: str | None = Form(None),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    item = db.query(CatalogoItem).filter(CatalogoItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item no encontrado")

    nombre_limpio = nombre.strip()
    if len(nombre_limpio) < 2:
        raise HTTPException(status_code=400, detail="Nombre de item invalido")

    duplicado = (
        db.query(CatalogoItem)
        .filter(CatalogoItem.nombre == nombre_limpio, CatalogoItem.id != item_id)
        .first()
    )
    if duplicado:
        return redirect_with_message("/admin/catalogo-items", "El item ya existe", "error")

    item.nombre = nombre_limpio
    item.activo = activo == "on"
    item.tipo_item = classify_catalog_item_type(nombre_limpio)
    item.permite_decimal = catalog_item_allows_decimal(nombre_limpio)
    db.commit()
    return redirect_with_message("/admin/catalogo-items", "Item actualizado", "success")


@app.post("/admin/catalogo-items/{item_id}/eliminar")
def admin_catalogo_item_eliminar(
    item_id: int,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_admin(current_user)
    item = db.query(CatalogoItem).filter(CatalogoItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item no encontrado")
    db.delete(item)
    db.commit()
    return redirect_with_message("/admin/catalogo-items", "Item eliminado", "warning")


@app.get("/api/requisiciones/{req_id}")
def detalle_requisicion(req_id: int, current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)):
    detail_options = [
        joinedload(Requisicion.items),
        joinedload(Requisicion.solicitante),
        joinedload(Requisicion.aprobador),
        joinedload(Requisicion.rechazador),
        joinedload(Requisicion.preparador),
        joinedload(Requisicion.entregador),
        joinedload(Requisicion.recibido_por),
        joinedload(Requisicion.receptor_designado),
        joinedload(Requisicion.liquidator),
        joinedload(Requisicion.prokey_ref_editor),
    ]
    maybe_prokey_liquidator = safe_joinedload(Requisicion, "prokey_liquidator")
    if maybe_prokey_liquidator is not None:
        detail_options.append(maybe_prokey_liquidator)

    req = (
        db.query(Requisicion)
        .options(*detail_options)
        .filter(Requisicion.id == req_id)
        .first()
    )
    if not req:
        raise HTTPException(status_code=404, detail="No existe")

    if not can_view_requisicion(req, current_user):
        raise HTTPException(status_code=403, detail="No autorizado")

    timeline: list[dict[str, object]] = []
    if req.created_at:
        timeline.append(
            {
                "evento": "Requisicion creada",
                "actor": req.solicitante.nombre if req.solicitante else None,
                "fecha_hora": req.created_at,
            }
        )
    if req.approved_at:
        timeline.append(
            {
                "evento": "Requisicion aprobada",
                "actor": req.aprobador.nombre if req.aprobador else None,
                "fecha_hora": req.approved_at,
            }
        )
    if req.prepared_at:
        timeline.append(
            {
                "evento": "Requisicion preparada",
                "actor": req.preparador.nombre if req.preparador else None,
                "fecha_hora": req.prepared_at,
            }
        )
    if req.rejected_at:
        timeline.append(
            {
                "evento": "Requisicion rechazada",
                "actor": req.rechazador.nombre if req.rechazador else None,
                "fecha_hora": req.rejected_at,
            }
        )
    if req.delivered_at:
        timeline.append(
            {
                "evento": "Entrega de bodega",
                "actor": req.entregador.nombre if req.entregador else None,
                "fecha_hora": req.delivered_at,
            }
        )
        if req.recibido_por and req.recibido_at:
            timeline.append(
                {
                    "evento": "Recibido con firma",
                    "actor": req.recibido_por.nombre,
                    "fecha_hora": req.recibido_at,
                }
            )
        elif req.delivered_to:
            timeline.append(
                {
                    "evento": "Recibido",
                    "actor": req.delivered_to,
                    "fecha_hora": req.delivered_at,
                }
            )
    if req.liquidated_at:
        timeline.append(
            {
                "evento": "Requisicion liquidada",
                "actor": req.liquidator.nombre if req.liquidator else None,
                "fecha_hora": req.liquidated_at,
            }
        )
    prokey_liquidator = getattr(req, "prokey_liquidator", None)
    prokey_liquidada_at = getattr(req, "prokey_liquidada_at", None)
    if prokey_liquidada_at:
        timeline.append(
            {
                "evento": "Liquidada en Prokey",
                "actor": prokey_liquidator.nombre if prokey_liquidator else None,
                "fecha_hora": prokey_liquidada_at,
            }
        )
    prokey_ref_editor = getattr(req, "prokey_ref_editor", None)
    prokey_ref_actualizada_at = getattr(req, "prokey_ref_actualizada_at", None)
    if req.prokey_ref and prokey_ref_actualizada_at:
        timeline.append(
            {
                "evento": "Referencia Prokey registrada",
                "actor": prokey_ref_editor.nombre if prokey_ref_editor else None,
                "fecha_hora": prokey_ref_actualizada_at,
            }
        )

    items_payload: list[dict[str, object]] = []
    for item in req.items:
        item_data: dict[str, object] = {
            "descripcion": item.descripcion,
            "cantidad": item.cantidad,
            "cantidad_entregada": item.cantidad_entregada,
            "unidad": item.unidad,
            "contexto_operacion": normalize_contexto_operacion(item.contexto_operacion),
            "es_demo": bool(item.es_demo),
        }
        if req.estado in ("liquidada", "liquidada_en_prokey"):
            mode = (item.liquidation_mode or "RETORNABLE").upper()
            if mode not in ("RETORNABLE", "CONSUMIBLE"):
                mode = "RETORNABLE"
            qty_returned = item.qty_returned_to_warehouse or 0
            qty_used = item.qty_used or 0
            qty_not_used = item.qty_left_at_client or 0
            delivered = item.cantidad_entregada or 0
            contexto_operacion = normalize_contexto_operacion(item.contexto_operacion)
            expected_return = calcular_retorno_esperado(mode, qty_used, qty_not_used, contexto_operacion)
            difference = expected_return - qty_returned
            pk_ingreso_qty = calcular_ingreso_pk_bodega(
                mode,
                delivered,
                qty_used,
                qty_returned,
                contexto_operacion,
            )

            parsed_alerts: list[dict[str, object]] = []
            if item.liquidation_alerts:
                try:
                    parsed = json.loads(item.liquidation_alerts)
                    if isinstance(parsed, list):
                        parsed_alerts = parsed
                except (json.JSONDecodeError, TypeError, ValueError):
                    parsed_alerts = []

            item_data.update(
                {
                    "mode": mode,
                    "used": qty_used,
                    "not_used": qty_not_used,
                    "returned": qty_returned,
                    "delivered": delivered,
                    "expected_return": expected_return,
                    "difference": difference,
                    "qty_returned_to_warehouse": qty_returned,
                    "qty_used": qty_used,
                    "qty_left_at_client": qty_not_used,
                    "item_liquidation_note": normalize_optional_text(item.item_liquidation_note),
                    "liquidation_alerts": parsed_alerts,
                    "qty_ocupo": qty_used + qty_not_used,
                    "pk_ingreso_qty": pk_ingreso_qty,
                    "delta": difference,
                    "contexto_operacion": contexto_operacion,
                }
            )
        items_payload.append(item_data)

    return {
        "id": req.id,
        "folio": req.folio,
        "solicitante": req.solicitante.nombre if req.solicitante else None,
        "departamento": req.departamento,
        "cliente_codigo": req.cliente_codigo,
        "cliente_nombre": req.cliente_nombre,
        "cliente_ruta_principal": req.cliente_ruta_principal,
        "estado": req.estado,
        "motivo_requisicion": req.motivo_requisicion,
        "justificacion": req.justificacion,
        "created_at": req.created_at,
        "approved_by": req.aprobador.nombre if req.aprobador else None,
        "approval_comment": req.approval_comment,
        "prepared_by": req.preparador.nombre if req.preparador else None,
        "prepared_at": req.prepared_at,
        "rejected_by": req.rechazador.nombre if req.rechazador else None,
        "rejection_comment": req.rejection_comment,
        "delivered_by": req.entregador.nombre if req.entregador else None,
        "delivered_to": req.delivered_to,
        "recibido_por": req.recibido_por.nombre if req.recibido_por else None,
        "recibido_por_detalle": (
            {
                "id": req.recibido_por.id,
                "nombre": req.recibido_por.nombre,
                "rol": req.recibido_por.rol,
            }
            if req.recibido_por
            else None
        ),
        "recibido_at": req.recibido_at,
        "receptor_designado": (
            {
                "id": req.receptor_designado.id,
                "nombre": req.receptor_designado.nombre,
                "rol": req.receptor_designado.rol,
            }
            if req.receptor_designado
            else None
        ),
        "delivery_result": req.delivery_result,
        "delivery_comment": req.delivery_comment,
        "rejection_reason": req.rejection_reason,
        "approved_at": req.approved_at,
        "prepared_by": req.preparador.nombre if req.preparador else None,
        "prepared_at": req.prepared_at,
        "rejected_at": req.rejected_at,
        "delivered_at": req.delivered_at,
        "prokey_ref": req.prokey_ref,
        "prokey_pending": not bool(req.prokey_ref),
        "prokey_ref_actualizada_at": prokey_ref_actualizada_at,
        "prokey_ref_actualizada_por_nombre": prokey_ref_editor.nombre if prokey_ref_editor else None,
        "prokey_ref_actualizada_por_rol": prokey_ref_editor.rol if prokey_ref_editor else None,
        "puede_editar_prokey_ref": puede_editar_prokey_ref(req, current_user),
        "liquidation_comment": normalize_optional_text(req.liquidation_comment),
        "liquidated_by_name": req.liquidator.nombre if req.liquidator else None,
        "liquidated_at": req.liquidated_at,
        "prokey_liquidada_at": prokey_liquidada_at,
        "prokey_liquidado_por_nombre": prokey_liquidator.nombre if prokey_liquidator else None,
        "pdf_url": (
            f"/requisiciones/{req.id}/pdf"
            if req.estado in ("aprobada", "preparado", "entregada", "liquidada", "liquidada_en_prokey")
            else None
        ),
        "timeline": timeline,
        "items": items_payload,
    }


@app.get("/requisiciones/{req_id}/pdf")
def descargar_pdf(req_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    from app.pdf_generator import generate_requisicion_pdf

    req = (
        db.query(Requisicion)
        .options(
            joinedload(Requisicion.items),
            joinedload(Requisicion.solicitante),
            joinedload(Requisicion.aprobador),
            joinedload(Requisicion.preparador),
            joinedload(Requisicion.entregador),
            joinedload(Requisicion.recibido_por),
            joinedload(Requisicion.receptor_designado),
            joinedload(Requisicion.liquidator),
        )
        .filter(Requisicion.id == req_id)
        .first()
    )
    if not req:
        raise HTTPException(status_code=404, detail="Requisición no encontrada")
    if not can_view_requisicion(req, current_user):
        raise HTTPException(status_code=403, detail="No autorizado")
    if req.estado not in ("aprobada", "preparado", "entregada", "liquidada", "liquidada_en_prokey"):
        raise HTTPException(status_code=403, detail="PDF disponible solo desde requisiciones aprobadas")

    items_data = []
    for item in req.items:
        alerts = item.liquidation_alerts
        if isinstance(alerts, str):
            try:
                alerts = json.loads(alerts)
            except Exception:
                alerts = []
        alert_types = []
        if isinstance(alerts, list):
            for alert in alerts:
                if isinstance(alert, dict):
                    alert_type = str(alert.get("type") or "").strip()
                    if alert_type:
                        alert_types.append(alert_type)
                elif alert:
                    alert_types.append(str(alert))
        items_data.append(
            {
                "descripcion": item.descripcion,
                "cantidad_solicitada": item.cantidad,
                "cantidad_entregada": item.cantidad_entregada,
                "cantidad_usada": item.qty_used,
                "cantidad_no_usada": item.qty_left_at_client,
                "cantidad_retorna": item.qty_returned_to_warehouse,
                "liquidation_mode": item.liquidation_mode,
                "contexto_operacion": item.contexto_operacion,
                "es_demo": bool(item.es_demo),
                "prokey_ref": req.prokey_ref,
                "pk_ingreso_qty": calcular_ingreso_pk_bodega(
                    item.liquidation_mode,
                    item.cantidad_entregada or 0,
                    item.qty_used or 0,
                    item.qty_returned_to_warehouse or 0,
                    item.contexto_operacion,
                ),
                "liquidation_alerts": alert_types,
                "nota_liquidacion": item.item_liquidation_note,
            }
        )

    req_data = {
        "id": req.id,
        "folio": req.folio or f"REQ-{req.id:04d}",
        "estado": req.estado,
        "created_at": str(req.created_at) if req.created_at else None,
        "approved_at": str(req.approved_at) if req.approved_at else None,
        "prepared_at": str(req.prepared_at) if req.prepared_at else None,
        "delivered_at": str(req.delivered_at) if req.delivered_at else None,
        "recibido_at": str(req.recibido_at) if req.recibido_at else None,
        "liquidated_at": str(req.liquidated_at) if req.liquidated_at else None,
        "cliente": req.cliente_nombre,
        "codigo_cliente": req.cliente_codigo,
        "ruta": req.cliente_ruta_principal,
        "solicitante_nombre": req.solicitante.nombre if req.solicitante else None,
        "receptor_designado_nombre": req.receptor_designado.nombre if req.receptor_designado else None,
        "receptor_designado_rol": req.receptor_designado.rol if req.receptor_designado else None,
        "aprobador_nombre": req.aprobador.nombre if req.aprobador else None,
        "preparador_nombre": req.preparador.nombre if req.preparador else None,
        "jefe_bodega_nombre": req.entregador.nombre if req.entregador else None,
        "liquidado_por_nombre": req.liquidator.nombre if req.liquidator else None,
        "recibido_por_nombre": req.recibido_por.nombre if req.recibido_por else None,
        "prokey_ref": req.prokey_ref,
        "justificacion": req.justificacion,
        "comentario_liquidacion": req.liquidation_comment,
        "items": items_data,
    }

    pdf_bytes = generate_requisicion_pdf(req_data)
    folio = req.folio or f"REQ-{req.id:04d}"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="requisicion_{folio}.pdf"'},
    )
